"""Export engine for CSV, Excel, and PDF reports."""

from __future__ import annotations

import csv
import logging
from datetime import datetime
from pathlib import Path
from typing import Any

logger = logging.getLogger(__name__)

OUTPUT_DIR = Path.home() / ".copilot-pulse" / "exports"


class ExportEngine:
    """Export Copilot metrics reports in various formats."""

    def __init__(self) -> None:
        OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    def export(
        self,
        fmt: str,
        title: str,
        sections: list[dict[str, Any]],
        filename: str = "copilot_report",
    ) -> Path:
        """Export a report in the specified format.

        Args:
            fmt: Export format ('csv', 'excel', 'pdf').
            title: Report title.
            sections: Report sections with data.
            filename: Output filename without extension.

        Returns:
            Path to the generated file.
        """
        match fmt:
            case "csv":
                return self._export_csv(title, sections, filename)
            case "excel":
                return self._export_excel(title, sections, filename)
            case "pdf":
                return self._export_pdf(title, sections, filename)
            case _:
                raise ValueError(f"Unsupported format: {fmt}")

    def _export_csv(
        self, title: str, sections: list[dict[str, Any]], filename: str
    ) -> Path:
        """Export as CSV file."""
        file_path = OUTPUT_DIR / f"{filename}.csv"

        with open(file_path, "w", newline="", encoding="utf-8") as f:
            writer = csv.writer(f)
            writer.writerow([f"# {title}"])
            writer.writerow([f"# Generated: {datetime.now().isoformat()}"])
            writer.writerow([])

            for section in sections:
                if isinstance(section, str):
                    section = {"title": "", "data": section}
                section_title = section.get("title", "")
                headers = section.get("headers", [])
                rows = section.get("rows", [])
                data = section.get("data", {})

                writer.writerow([f"## {section_title}"])

                if headers and rows:
                    writer.writerow(headers)
                    for row in rows:
                        writer.writerow(row)
                elif isinstance(data, dict):
                    for key, value in data.items():
                        writer.writerow([key, value])
                elif data:
                    writer.writerow([str(data)])

                writer.writerow([])

        logger.info("CSV exported to %s", file_path)
        return file_path

    def _export_excel(
        self, title: str, sections: list[dict[str, Any]], filename: str
    ) -> Path:
        """Export as Excel file with formatting."""
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font, PatternFill
        from openpyxl.utils import get_column_letter

        file_path = OUTPUT_DIR / f"{filename}.xlsx"
        wb = Workbook()

        # Summary sheet
        summary = wb.active
        summary.title = "Summary"
        summary["A1"] = title
        summary["A1"].font = Font(size=16, bold=True)
        summary["A2"] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
        summary["A2"].font = Font(size=10, italic=True, color="666666")

        row_offset = 4
        for section in sections:
            if isinstance(section, str):
                section = {"title": "", "data": section}
            section_title = section.get("title", "Data")
            headers = section.get("headers", [])
            rows = section.get("rows", [])
            data = section.get("data", {})

            summary.cell(row=row_offset, column=1, value=section_title).font = Font(
                bold=True, size=12
            )
            row_offset += 1

            if headers and rows:
                # Create a dedicated sheet for detailed data
                ws = wb.create_sheet(title=section_title[:31])
                header_fill = PatternFill(start_color="1f6feb", end_color="1f6feb", fill_type="solid")
                header_font = Font(bold=True, color="FFFFFF")

                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal="center")
                    ws.column_dimensions[get_column_letter(col_idx)].width = max(len(str(header)) + 4, 12)

                for row_idx, row in enumerate(rows, 2):
                    for col_idx, val in enumerate(row, 1):
                        ws.cell(row=row_idx, column=col_idx, value=val)

                summary.cell(row=row_offset, column=1, value=f"→ See '{section_title}' sheet")
                row_offset += 1

            elif isinstance(data, dict):
                for key, value in data.items():
                    summary.cell(row=row_offset, column=1, value=str(key))
                    summary.cell(row=row_offset, column=2, value=str(value))
                    row_offset += 1
            elif data:
                summary.cell(row=row_offset, column=1, value=str(data))
                row_offset += 1

            row_offset += 1

        wb.save(str(file_path))
        logger.info("Excel exported to %s", file_path)
        return file_path

    def _export_pdf(
        self, title: str, sections: list[dict[str, Any]], filename: str
    ) -> Path:
        """Export as PDF report."""
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
        from reportlab.lib.units import mm
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )

        file_path = OUTPUT_DIR / f"{filename}.pdf"
        doc = SimpleDocTemplate(
            str(file_path),
            pagesize=A4,
            topMargin=20 * mm,
            bottomMargin=20 * mm,
        )

        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            "CustomTitle", parent=styles["Title"], fontSize=20, spaceAfter=12
        )
        subtitle_style = ParagraphStyle(
            "Subtitle", parent=styles["Normal"], fontSize=10, textColor=colors.grey
        )
        section_style = ParagraphStyle(
            "SectionTitle", parent=styles["Heading2"], fontSize=14, spaceBefore=16
        )

        elements = []
        elements.append(Paragraph(title, title_style))
        elements.append(
            Paragraph(
                f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}",
                subtitle_style,
            )
        )
        elements.append(Spacer(1, 12))

        for section in sections:
            if isinstance(section, str):
                section = {"title": "", "data": section}
            section_title = section.get("title", "")
            headers = section.get("headers", [])
            rows = section.get("rows", [])
            data = section.get("data", {})
            notes = section.get("notes", "")

            elements.append(Paragraph(section_title, section_style))

            if headers and rows:
                table_data = [headers] + [[str(v) for v in row] for row in rows]
                t = Table(table_data, repeatRows=1)
                t.setStyle(
                    TableStyle([
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f6feb")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, 0), 9),
                        ("FONTSIZE", (0, 1), (-1, -1), 8),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f6f8fa")]),
                        ("ALIGN", (1, 0), (-1, -1), "CENTER"),
                    ])
                )
                elements.append(t)

            elif isinstance(data, dict):
                kv_data = [[str(k), str(v)] for k, v in data.items()]
                t = Table(kv_data)
                t.setStyle(
                    TableStyle([
                        ("FONTNAME", (0, 0), (0, -1), "Helvetica-Bold"),
                        ("FONTSIZE", (0, 0), (-1, -1), 9),
                        ("GRID", (0, 0), (-1, -1), 0.5, colors.lightgrey),
                    ])
                )
                elements.append(t)
            elif data:
                elements.append(Paragraph(str(data), styles["Normal"]))

            if notes:
                elements.append(Spacer(1, 6))
                elements.append(Paragraph(notes, styles["Normal"]))

            elements.append(Spacer(1, 12))

        doc.build(elements)
        logger.info("PDF exported to %s", file_path)
        return file_path
