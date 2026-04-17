"""FastAPI web dashboard with HTMX and WebSocket chat."""

from __future__ import annotations

import asyncio
import io
import json
import logging
import shutil
import tempfile
import time
import uuid
from pathlib import Path
from typing import Any

from fastapi import FastAPI, File, Form, UploadFile, WebSocket, WebSocketDisconnect
from fastapi.responses import HTMLResponse, JSONResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from starlette.requests import Request

from starlette.responses import RedirectResponse

from ..agent.providers import create_provider_from_settings
from ..config import AppConfig, load_settings_file, save_settings_file
from ..github_client.auth import build_github_auth
from ..github_client.base_client import GitHubBaseClient
from ..github_client.models import ReportDownloadResponse
from ..github_client.usage_metrics_api import _unwrap_day_totals
from ..orgdata.database import OrgDatabase
from ..web.i18n import get_translations

logger = logging.getLogger(__name__)

# Cache TTL defaults (seconds) — overridden at runtime from AppConfig / .env.
# See CACHE_TTL_28D, CACHE_TTL_WEEKLY, CACHE_TTL_SEATS in .env.
_TTL_28D = 1800       # 30 min — 28-day org/user report
_TTL_WEEKLY = 3600    # 60 min — historical Wednesday snapshots never change once written
_TTL_SEATS = 1800     # 30 min — seat assignments change rarely

# In-memory cache for raw 28-day records (avoids re-fetching per chart).
_raw_cache: dict[str, Any] = {"data": None, "ts": 0.0}
_user_cache: dict[str, Any] = {"data": None, "ts": 0.0}
_seat_cache: dict[str, Any] = {"data": None, "ts": 0.0}

# Temporary upload sessions for the two-step org import flow.
# Maps session_id → {"path": str, "created_at": float}
_upload_sessions: dict[str, dict[str, Any]] = {}
_UPLOAD_SESSION_TTL = 1800  # 30 minutes


def _cleanup_upload_sessions() -> None:
    cutoff = time.time() - _UPLOAD_SESSION_TTL
    expired = [sid for sid, s in _upload_sessions.items() if s["created_at"] < cutoff]
    for sid in expired:
        try:
            Path(_upload_sessions[sid]["path"]).unlink(missing_ok=True)
        except Exception:
            pass
        del _upload_sessions[sid]

WEB_DIR = Path(__file__).resolve().parent.parent / "web"
TEMPLATES_DIR = WEB_DIR / "templates"
STATIC_DIR = WEB_DIR / "static"


def _latest_metric_record(records: list[dict[str, Any]]) -> dict[str, Any]:
    """Return the most recent metric record from a raw metrics payload."""
    if not records:
        return {}
    return max(records, key=lambda r: r.get("day") or r.get("date", ""))


def _sum_agent_edits(records: list[dict[str, Any]]) -> int:
    """Sum agent_edit activity across raw metric records."""
    total = 0
    for rec in records:
        for feat in rec.get("totals_by_feature", []):
            if feat.get("feature") == "agent_edit":
                total += feat.get("code_generation_activity_count", 0)
    return total


def _resolve_active_user_count(records: list[dict[str, Any]], *, is_filtered: bool) -> int:
    """Return the active-user denominator used by productivity KPIs."""
    if not records:
        return 0
    if is_filtered:
        return len({r.get("user_login") for r in records if r.get("user_login")})

    latest = _latest_metric_record(records)
    return (
        latest.get("monthly_active_users")
        or latest.get("total_engaged_users")
        or latest.get("total_active_users")
        or 0
    )


def _compute_productivity_metrics(
    *, agent_edits: int, active_users: int, total_licenses: int
) -> dict[str, Any]:
    """Compute the productivity KPI values for the dashboard."""
    efficacy = agent_edits / active_users if active_users else None
    real_adoption = agent_edits / total_licenses if total_licenses else None
    return {
        "agent_edits": agent_edits,
        "active_users": active_users,
        "total_licenses": total_licenses,
        "agent_efficacy": efficacy,
        "real_adoption": real_adoption,
        "efficacy_badge": "Advanced" if efficacy is not None and efficacy > 50 else None,
    }


# Weekly chart bars come from 1-day API snapshots (one Wednesday per week),
# so thresholds are calibrated for a single day: 28D / 28 days.
_CANONICAL_THRESHOLDS: list[float] = [0.36, 1.79, 3.57]       # daily  (28D / 28)
_CANONICAL_THRESHOLDS_28D: list[float] = [10.0, 50.0, 100.0]  # 28-day
_THRESHOLD_COLORS: list[str] = ["#f85149", "#d29922", "#3fb950", "#58a6ff"]


def _agent_edit_weekly_color_with_thresholds(ratio: float, thresholds: list[float]) -> str:
    """Return the UI color for a weekly agent-edits ratio given explicit thresholds."""
    t1, t2, t3 = thresholds
    if ratio < t1:
        return _THRESHOLD_COLORS[0]
    if ratio <= t2:
        return _THRESHOLD_COLORS[1]
    if ratio <= t3:
        return _THRESHOLD_COLORS[2]
    return _THRESHOLD_COLORS[3]


def _agent_edit_weekly_color(ratio: float) -> str:
    """Return the UI color for a weekly agent-edits-per-seat value (canonical thresholds)."""
    return _agent_edit_weekly_color_with_thresholds(ratio, _CANONICAL_THRESHOLDS)


def _compute_active_rate(active_users_per_week: list[int], total_seats: int) -> float:
    """Return avg(active_users) / total_seats, clamped to [0.01, 1.0].

    Falls back to 1.0 (no scaling) when data is unavailable, so callers
    can always divide by this value without special-casing.
    """
    if not total_seats or not active_users_per_week:
        return 1.0
    nonzero = [u for u in active_users_per_week if u > 0]
    if not nonzero:
        return 1.0
    avg_active = sum(nonzero) / len(nonzero)
    return min(max(avg_active / total_seats, 0.01), 1.0)


def _build_scale_legend_html(
    thresholds: list[float],
    labels: dict[str, str],
) -> str:
    """Return an HTML .productivity-scale block with the given threshold values."""

    def _fmt(v: float) -> str:
        return str(int(v)) if v == int(v) else f"{v:.1f}"

    t1, t2, t3 = thresholds
    bands = [f"&lt; {_fmt(t1)}", f"{_fmt(t1)}–{_fmt(t2)}", f"{_fmt(t2)}–{_fmt(t3)}", f"&gt; {_fmt(t3)}"]
    items = zip(bands, _THRESHOLD_COLORS, [
        labels.get("product_scale_cautious", "Cautious / Legacy"),
        labels.get("product_scale_standard", "Standard Adopters"),
        labels.get("product_scale_advanced", "Advanced"),
        labels.get("product_scale_agent_first", "Agent-First / Power Users"),
    ])
    html = '<div class="productivity-scale">'
    for band, color, label in items:
        html += (
            f'<div class="productivity-scale-item">'
            f'<span class="agent-dot" style="background:{color}"></span>'
            f'<span>{band} — {label}</span>'
            f'</div>'
        )
    html += "</div>"
    return html


def create_app(config: AppConfig, refresh_maturity_cache: bool = False) -> FastAPI:
    """Create and configure the FastAPI dashboard application.

    Args:
        config: Application configuration.
        refresh_maturity_cache: If True, invalidate any persisted maturity cache
            entries on startup so the first request triggers a fresh computation.

    Returns:
        Configured FastAPI app.
    """
    app = FastAPI(title="Copilot Pulse Dashboard", version="0.1.0")
    templates = Jinja2Templates(directory=str(TEMPLATES_DIR))

    if STATIC_DIR.exists():
        app.mount("/static", StaticFiles(directory=str(STATIC_DIR)), name="static")

    # Store config and orchestrator reference on app state
    app.state.config = config
    app.state.orchestrator = None  # Set externally before startup

    import time as _time
    # Resolve TTLs from config (env-configurable); fall back to module defaults.
    _ttl_28d = config.cache_ttl_28d
    _ttl_weekly = config.cache_ttl_weekly
    _ttl_seats = config.cache_ttl_seats

    # ── Persistent maturity cache (24-hour TTL, survives restarts) ──
    from ..cache.store import CacheStore as _CacheStore
    _MATURITY_TTL_H = 24
    _maturity_store = _CacheStore(ttl_hours=_MATURITY_TTL_H)
    if refresh_maturity_cache:
        removed = _maturity_store.delete_prefix("maturity:")
        logger.info("Maturity cache refreshed: %d entries removed", removed)

    # ── Org-structure filter mapping ────────────────────────────
    # github_login → {level4..level8}; populated from the SQLite org DB.
    _org_map: dict[str, dict[str, str]] = {}
    _org_levels: dict[str, list[str]] = {"4": [], "5": [], "6": [], "7": [], "8": []}

    def _reload_org_map() -> None:
        """Rebuild _org_map and _org_levels from the SQLite org database."""
        _org_map.clear()
        for lvl in _org_levels:
            _org_levels[lvl] = []

        db = _get_orgdb()
        rows = db._conn.execute(
            "SELECT github_id, sup_org_level_4, sup_org_level_5, "
            "sup_org_level_6, sup_org_level_7, sup_org_level_8 "
            "FROM employees WHERE github_id IS NOT NULL AND github_id != ''"
        ).fetchall()

        if not rows:
            logger.info("Org filter map: no employees with GitHub ID in DB — filters disabled")
            return

        sets: dict[str, set[str]] = {"4": set(), "5": set(), "6": set(), "7": set(), "8": set()}
        for row in rows:
            gh = str(row[0]).strip().lower()
            l4 = str(row[1] or "").strip()
            l5 = str(row[2] or "").strip()
            l6 = str(row[3] or "").strip()
            l7 = str(row[4] or "").strip()
            l8 = str(row[5] or "").strip()
            _org_map[gh] = {"4": l4, "5": l5, "6": l6, "7": l7, "8": l8}
            for lvl, val in (("4", l4), ("5", l5), ("6", l6), ("7", l7), ("8", l8)):
                if val:
                    sets[lvl].add(val)

        for lvl in ("4", "5", "6", "7", "8"):
            _org_levels[lvl] = sorted(sets[lvl])

        logger.info(
            "Org filter map loaded from DB: %d users, L6=%d L7=%d L8=%d values",
            len(_org_map), len(_org_levels["6"]), len(_org_levels["7"]), len(_org_levels["8"]),
        )

    def _filter_logins(level: str, value: str) -> set[str]:
        """Return the set of github logins matching a given org filter."""
        return {
            login for login, levels in _org_map.items()
            if levels.get(level) == value
        }

    # ── Data fetching with caching ──────────────────────────────

    async def _get_raw_org_metrics() -> list[dict[str, Any]]:
        """Fetch raw (unparsed) 28-day org metrics with caching."""
        now = _time.time()
        if _raw_cache["data"] is not None and now - _raw_cache["ts"] < _ttl_28d:
            return _raw_cache["data"]

        auth = build_github_auth(config)
        client = GitHubBaseClient(auth)
        try:
            resp = await client.get(
                f"/orgs/{config.github_org}/copilot/metrics/reports/organization-28-day/latest"
            )
            dr = ReportDownloadResponse(**resp.json())
            raw_records: list[dict[str, Any]] = []
            for url in dr.download_links:
                raw_records.extend(await client.download_ndjson(url))
            unwrapped = _unwrap_day_totals(raw_records)
            _raw_cache["data"] = unwrapped
            _raw_cache["ts"] = now
            return unwrapped
        finally:
            await client.close()

    async def _get_raw_user_metrics() -> list[dict[str, Any]]:
        """Fetch raw per-user 28-day metrics with caching."""
        now = _time.time()
        if _user_cache["data"] is not None and now - _user_cache["ts"] < _ttl_28d:
            return _user_cache["data"]

        auth = build_github_auth(config)
        client = GitHubBaseClient(auth)
        try:
            resp = await client.get(
                f"/orgs/{config.github_org}/copilot/metrics/reports/users-28-day/latest"
            )
            dr = ReportDownloadResponse(**resp.json())
            raw_records: list[dict[str, Any]] = []
            for url in dr.download_links:
                raw_records.extend(await client.download_ndjson(url))
            unwrapped = _unwrap_day_totals(raw_records)
            _user_cache["data"] = unwrapped
            _user_cache["ts"] = now
            return unwrapped
        finally:
            await client.close()

    def _parse_filter(request: Request) -> tuple[str | None, str | None]:
        """Extract org filter from query params.  Returns (level, value) or (None, None)."""
        level = request.query_params.get("filter_level")
        value = request.query_params.get("filter_value")
        if level and value:
            return level, value
        return None, None

    _ATTR_FILTER_DIMS: frozenset[str] = frozenset({
        "job_profile", "job_family", "business_title",
        "management_level", "is_manager", "age_range",
    })

    _AGE_RANGE_CASE_SIMPLE = """
        CASE
            WHEN age IS NULL THEN 'Unknown'
            WHEN age < 25   THEN '<25'
            WHEN age < 30   THEN '25-29'
            WHEN age < 35   THEN '30-34'
            WHEN age < 40   THEN '35-39'
            WHEN age < 45   THEN '40-44'
            WHEN age < 50   THEN '45-49'
            WHEN age < 55   THEN '50-54'
            WHEN age < 60   THEN '55-59'
            ELSE '60+'
        END
    """

    def _parse_attr_filters(request: Request) -> dict[str, list[str]]:
        """Extract attr_* query params. Returns {dim: [values...]}."""
        attrs: dict[str, list[str]] = {}
        for key, val in request.query_params.multi_items():
            if key.startswith("attr_") and val:
                dim = key[5:]
                if dim in _ATTR_FILTER_DIMS:
                    attrs.setdefault(dim, []).append(val)
        return attrs

    def _attr_filter_logins(attrs: dict[str, list[str]]) -> set[str] | None:
        """Query SQLite employees to get github_ids matching attribute filters."""
        if not attrs:
            return None
        db = _get_orgdb()
        where_clauses: list[str] = []
        params: list[Any] = []
        for dim, values in attrs.items():
            if not values:
                continue
            if dim == "age_range":
                ph = ",".join("?" * len(values))
                where_clauses.append(f"({_AGE_RANGE_CASE_SIMPLE}) IN ({ph})")
                params.extend(values)
            elif dim == "is_manager":
                mapped = [1 if v.lower() in ("true", "yes", "1") else 0 for v in values]
                ph = ",".join("?" * len(mapped))
                where_clauses.append(f"is_manager IN ({ph})")
                params.extend(mapped)
            elif dim in {"job_profile", "job_family", "business_title", "management_level"}:
                ph = ",".join("?" * len(values))
                where_clauses.append(f"LOWER(COALESCE({dim}, '')) IN ({ph})")
                params.extend(v.lower() for v in values)
        if not where_clauses:
            return None
        sql = (
            "SELECT github_id FROM employees "
            "WHERE github_id IS NOT NULL AND github_id != '' AND "
            + " AND ".join(where_clauses)
        )
        rows = db._conn.execute(sql, params).fetchall()
        return {str(r[0]).strip().lower() for r in rows}

    def _resolve_filter_logins(request: Request) -> set[str] | None:
        """Combine org hierarchy filter + attribute filters. Returns None if no filter active."""
        level, value = _parse_filter(request)
        org_set = _filter_logins(level, value) if level and value else None
        attr_set = _attr_filter_logins(_parse_attr_filters(request))
        if org_set is not None and attr_set is not None:
            return org_set & attr_set
        return org_set if org_set is not None else attr_set

    def _filter_cache_key(request: Request) -> str:
        """Stable cache key incorporating org + attribute filter params."""
        level, value = _parse_filter(request)
        attrs = _parse_attr_filters(request)
        parts = [f"{level}:{value}" if (level and value) else "__all__"]
        for dim in sorted(attrs.keys()):
            parts.append(f"{dim}={','.join(sorted(attrs[dim]))}")
        return "|".join(parts)

    async def _get_filtered_records(request: Request) -> tuple[list[dict[str, Any]], bool]:
        """Return records to use — either org-level or filtered user-level.

        Returns (records, is_filtered).
        """
        filter_logins = _resolve_filter_logins(request)
        if filter_logins is None:
            return await _get_raw_org_metrics(), False
        if not filter_logins:
            return [], True
        user_records = await _get_raw_user_metrics()
        filtered = [
            r for r in user_records
            if r.get("user_login", "").lower() in filter_logins
        ]
        return filtered, True

    async def _resolve_total_licenses(request: Request, *, is_filtered: bool) -> int:
        """Resolve the license denominator for adoption-style KPIs."""
        if is_filtered:
            fl = _resolve_filter_logins(request)
            return len(fl) if fl else 0

        now = _time.time()
        if _seat_cache["data"] is not None and now - _seat_cache["ts"] < _ttl_seats:
            return _seat_cache["data"]

        orch = app.state.orchestrator
        if not orch:
            return 0
        try:
            seat_result = await orch._tool_seat_info({"org": config.github_org})
            seat_info = seat_result.get("seat_info", {})
            total = seat_info.get("total_seats", 0) or 0
            _seat_cache["data"] = total
            _seat_cache["ts"] = now
            return total
        except Exception:
            return 0

    def _weekly_reference_dates() -> list[Any]:
        """Return 13 weekly reference points ending with the latest Wednesday."""
        from datetime import date as _date, timedelta as _td

        base = _date.today() - _td(days=2)
        days_since_wed = (base.weekday() - 2) % 7
        last_wed = base - _td(days=days_since_wed)
        return [last_wed - _td(weeks=i) for i in range(12, -1, -1)]

    _weekly_agent_cache: dict[str, dict[str, Any]] = {}
    # Per-key locks prevent multiple concurrent callers (e.g. 4 chart endpoints firing in
    # parallel on a cold cache) from each independently kicking off 13 API calls.
    _weekly_agent_locks: dict[str, asyncio.Lock] = {}

    async def _get_weekly_agent_series(request: Request) -> dict[str, Any]:
        """Fetch weekly agent-edit samples shared by multiple dashboard charts."""
        cache_key = _filter_cache_key(request)
        now = _time.time()
        cached = _weekly_agent_cache.get(cache_key)
        if cached and now - cached["ts"] < _ttl_weekly:
            return cached["data"]

        # Only one coroutine per cache key may run the 13 API calls at a time.
        if cache_key not in _weekly_agent_locks:
            _weekly_agent_locks[cache_key] = asyncio.Lock()
        async with _weekly_agent_locks[cache_key]:
            # Re-check after acquiring lock — another waiter may have populated it.
            cached = _weekly_agent_cache.get(cache_key)
            if cached and now - cached["ts"] < _ttl_weekly:
                return cached["data"]

            filter_logins = _resolve_filter_logins(request)
            if filter_logins is not None and not filter_logins:
                weeks = _weekly_reference_dates()
                empty = {
                    "labels": [f"W{w.isocalendar()[1]} ({w.strftime('%d/%m')})" for w in weeks],
                    "agent_edits": [0 for _ in weeks],
                    "active_users": [0 for _ in weeks],
                    "ratios": [0.0 for _ in weeks],
                }
                _weekly_agent_cache[cache_key] = {"data": empty, "ts": now}
                return empty

            use_user_endpoint = filter_logins is not None
            auth = build_github_auth(config)
            client = GitHubBaseClient(auth)
            weeks = _weekly_reference_dates()

            async def _fetch_day(day_str: str) -> list[dict[str, Any]] | dict[str, Any] | None:
                try:
                    endpoint = (
                        f"/orgs/{config.github_org}/copilot/metrics/reports/users-1-day"
                        if use_user_endpoint
                        else f"/orgs/{config.github_org}/copilot/metrics/reports/organization-1-day"
                    )
                    resp = await client.get(endpoint, params={"day": day_str})
                    dr = ReportDownloadResponse(**resp.json())
                    if not dr.download_links:
                        return None
                    raw = await client.download_ndjson(dr.download_links[0])
                    recs = _unwrap_day_totals(raw)
                    if use_user_endpoint:
                        return recs
                    return recs[0] if recs else None
                except Exception as exc:
                    logger.warning("weekly agent series: failed to fetch %s: %s", day_str, exc)
                    return None

            try:
                results = await asyncio.gather(*[_fetch_day(w.isoformat()) for w in weeks])

                labels: list[str] = []
                agent_edits_values: list[int] = []
                active_users_values: list[int] = []
                ratio_values: list[float] = []
                lines_accepted_values: list[int] = []

                for week_date, result in zip(weeks, results):
                    labels.append(f"W{week_date.isocalendar()[1]} ({week_date.strftime('%d/%m')})")
                    if result is None:
                        agent_edits_values.append(0)
                        active_users_values.append(0)
                        ratio_values.append(0.0)
                        lines_accepted_values.append(0)
                        continue

                    if use_user_endpoint:
                        filtered = [
                            r for r in result
                            if r.get("user_login", "").lower() in filter_logins
                        ]
                        agent_edits = _sum_agent_edits(filtered)
                        active_users = len({
                            r.get("user_login") for r in filtered if r.get("user_login")
                        })
                        lines_accepted = sum(
                            feat.get("lines_accepted", 0) or feat.get("code_acceptance_activity_count", 0)
                            for r in filtered
                            for feat in r.get("totals_by_feature", [])
                            if feat.get("feature") == "code_completion"
                        )
                    else:
                        rec = result
                        agent_edits = _sum_agent_edits([rec])
                        active_users = (
                            rec.get("monthly_active_users")
                            or rec.get("total_engaged_users")
                            or rec.get("total_active_users")
                            or 0
                        )
                        comp = rec.get("copilot_ide_code_completions") or {}
                        lines_accepted = (
                            comp.get("total_lines_accepted", 0)
                            or comp.get("total_code_acceptances", 0)
                        )

                    agent_edits_values.append(agent_edits)
                    active_users_values.append(active_users)
                    ratio_values.append(round(agent_edits / active_users, 1) if active_users else 0.0)
                    lines_accepted_values.append(lines_accepted)

                data = {
                    "labels": labels,
                    "agent_edits": agent_edits_values,
                    "active_users": active_users_values,
                    "ratios": ratio_values,
                    "lines_accepted": lines_accepted_values,
                }
                _weekly_agent_cache[cache_key] = {"data": data, "ts": now}
                return data
            finally:
                await client.close()

    async def _get_weekly_agent_ratio_payload(request: Request) -> dict[str, Any]:
        """Return weekly Agent Edits / Active User values with dynamically scaled thresholds."""
        series = await _get_weekly_agent_series(request)
        values = [round(v, 1) for v in series["ratios"]]
        average = round(sum(values) / len(values), 1) if values else 0.0

        level, value = _parse_filter(request)
        is_filtered = bool(level and value)
        total_seats = await _resolve_total_licenses(request, is_filtered=is_filtered)

        active_rate = _compute_active_rate(series["active_users"], total_seats)
        thresholds = [round(t / active_rate, 2) for t in _CANONICAL_THRESHOLDS]

        return {
            "labels": series["labels"],
            "values": values,
            "colors": [_agent_edit_weekly_color_with_thresholds(v, thresholds) for v in values],
            "average": average,
            "thresholds": thresholds,
            "active_rate": round(active_rate, 4),
        }

    def _lang(request: Request) -> str:
        """Read language preference from cookie (default: en)."""
        return request.cookies.get("lang", "en")

    def _ctx(request: Request, **extra: Any) -> dict[str, Any]:
        """Build common template context with translations."""
        lang = _lang(request)
        return {
            "request": request,
            "enterprise": config.github_enterprise,
            "org": config.github_org,
            "lang": lang,
            "t": get_translations(lang),
            **extra,
        }

    @app.get("/set-lang", response_class=HTMLResponse)
    async def set_lang(request: Request, lang: str = "en"):
        """Switch UI language and redirect back."""
        if lang not in ("en", "it"):
            lang = "en"
        referer = request.headers.get("referer", "/")
        response = RedirectResponse(url=referer, status_code=302)
        response.set_cookie("lang", lang, max_age=365 * 86400, samesite="lax")
        return response

    @app.get("/api/filter-options")
    async def api_filter_options():
        """Return distinct values + counts for each people-attribute filter dimension."""
        try:
            db = _get_orgdb()
            result: dict[str, list[dict]] = {}

            _DIM_COLS = [
                ("job_profile",      "job_profile"),
                ("job_family",       "job_family"),
                ("business_title",   "business_title"),
                ("management_level", "management_level"),
            ]
            for dim_key, col in _DIM_COLS:
                rows = db._conn.execute(
                    f"SELECT COALESCE({col}, '') AS v, COUNT(*) AS cnt "
                    f"FROM employees WHERE {col} IS NOT NULL AND {col} != '' "
                    f"GROUP BY v ORDER BY cnt DESC"
                ).fetchall()
                if rows:
                    result[dim_key] = [{"value": r[0], "count": r[1]} for r in rows]

            # is_manager (boolean)
            mgr_rows = db._conn.execute(
                "SELECT is_manager, COUNT(*) AS cnt FROM employees "
                "WHERE is_manager IS NOT NULL GROUP BY is_manager ORDER BY is_manager"
            ).fetchall()
            if mgr_rows:
                result["is_manager"] = [
                    {"value": "true" if r[0] else "false", "count": r[1]}
                    for r in mgr_rows
                ]

            # age_range (computed bucket)
            age_rows = db._conn.execute(
                f"SELECT ({_AGE_RANGE_CASE_SIMPLE}) AS ar, COUNT(*) AS cnt "
                f"FROM employees GROUP BY ar ORDER BY ar"
            ).fetchall()
            if age_rows:
                result["age_range"] = [
                    {"value": r[0], "count": r[1]} for r in age_rows if r[0] != "Unknown"
                ]

            return result
        except Exception as e:
            logger.error("filter-options error: %s", e)
            return {}

    @app.get("/api/org-filters")
    async def api_org_filters():
        """Return available org filter values for cascading dropdowns."""
        return {
            "enabled": bool(_org_map),
            "mapped_users": len(_org_map),
            "levels": {
                "4": _org_levels["4"],
                "5": _org_levels["5"],
                "6": _org_levels["6"],
                "7": _org_levels["7"],
                "8": _org_levels["8"],
            },
            "children": {
                "4": {
                    v4: sorted({
                        m["5"] for m in _org_map.values()
                        if m["4"] == v4 and m["5"]
                    })
                    for v4 in _org_levels["4"]
                },
                "5": {
                    v5: sorted({
                        m["6"] for m in _org_map.values()
                        if m["5"] == v5 and m["6"]
                    })
                    for v5 in _org_levels["5"]
                },
                "6": {
                    v6: sorted({
                        m["7"] for m in _org_map.values()
                        if m["6"] == v6 and m["7"]
                    })
                    for v6 in _org_levels["6"]
                },
                "7": {
                    v7: sorted({
                        m["8"] for m in _org_map.values()
                        if m["7"] == v7 and m["8"]
                    })
                    for v7 in _org_levels["7"]
                },
            },
        }

    @app.get("/api/data-as-of", response_class=HTMLResponse)
    async def api_data_as_of(request: Request):
        """Return an HTML badge showing the most recent date with actual data."""
        lang = _lang(request)
        t = get_translations(lang)
        try:
            org_records = await _get_raw_org_metrics()
            latest = max(
                (r.get("date") or r.get("day", "") for r in org_records), default=""
            )
            if latest:
                label = t.get("dash_data_as_of", "Data as of")
                return HTMLResponse(
                    f'<span class="data-as-of-badge">{label}: <strong>{latest}</strong></span>'
                )
        except Exception:
            pass
        return HTMLResponse("")

    @app.get("/", response_class=HTMLResponse)
    async def dashboard(request: Request):
        """Main dashboard page with KPI cards and charts."""
        return templates.TemplateResponse(
            "dashboard.html",
            _ctx(request, title="Copilot Pulse Dashboard",
                 has_org_filters=bool(_org_map)),
        )

    @app.get("/chat", response_class=HTMLResponse)
    async def chat_page(request: Request):
        """Chat interface page."""
        return templates.TemplateResponse(
            "dashboard.html",
            _ctx(request, title="Copilot Pulse Chat", active_tab="chat"),
        )

    @app.get("/api/metrics", response_class=HTMLResponse)
    async def api_metrics(request: Request):
        """Return KPI cards as an HTML fragment for HTMX swap."""
        lang = _lang(request)
        t = get_translations(lang)
        filter_logins = _resolve_filter_logins(request)

        if filter_logins is not None:
            # Filtered mode — compute KPIs from user-level data
            try:
                logins = filter_logins
                # Anchor latest_date from org-level records: they always have one entry
                # per day for all 28 days, so max(date) is reliably the most recent day
                # with data (~today-2) regardless of how sparse user activity is.
                org_records = await _get_raw_org_metrics()
                latest_date = max(
                    (r.get("date") or r.get("day", "") for r in org_records), default=""
                )
                user_records = await _get_raw_user_metrics()
                filtered = [r for r in user_records if r.get("user_login", "").lower() in logins]
                if not filtered:
                    return HTMLResponse(
                        f'<div class="kpi-card"><div class="kpi-value">0</div>'
                        f'<div class="kpi-label">{t.get("dash_active_users", "Active Users")}</div></div>'
                    )
                # Engaged = unique users across the full 28-day window
                # Active  = unique users present on the latest available day
                active = len({
                    r.get("user_login") for r in filtered
                    if r.get("user_login") and (r.get("date") or r.get("day", "")) == latest_date
                })
                engaged = len({r.get("user_login") for r in filtered if r.get("user_login")})
                sugg, acc = 0, 0
                for rec in filtered:
                    for feat in rec.get("totals_by_feature", []):
                        sugg += feat.get("code_generation_activity_count", 0)
                        acc += feat.get("code_acceptance_activity_count", 0)
                rate = f"{acc / sugg * 100:.1f}%" if sugg else "N/A"
                seats_label = str(len(logins))
            except Exception as e:
                logger.error("API metrics (filtered) error: %s", e)
                return HTMLResponse(
                    f'<div class="kpi-card"><div class="kpi-value">!</div>'
                    f'<div class="kpi-label">{e}</div></div>'
                )
        else:
            # Unfiltered mode — count unique users from user-level data (same 28-day
            # window as the filtered path) so the numbers are directly comparable.
            # Org-level aggregate is only used for suggestions/acceptances.
            orch = app.state.orchestrator
            if not orch:
                return HTMLResponse(
                    f'<div class="kpi-card"><div class="kpi-value">!</div>'
                    f'<div class="kpi-label">Agent not initialized</div></div>'
                )
            try:
                # Anchor latest_date from org-level records (complete daily series)
                org_records = await _get_raw_org_metrics()
                latest_date = max(
                    (r.get("date") or r.get("day", "") for r in org_records), default=""
                )
                user_records = await _get_raw_user_metrics()
                if not user_records:
                    return HTMLResponse(
                        f'<div class="kpi-card"><div class="kpi-value">0</div>'
                        f'<div class="kpi-label">{t.get("dash_active_users", "Active Users")}</div></div>'
                    )
                # Engaged = unique users across the full 28-day window
                # Active  = unique users present on the latest available day
                active = len({
                    r.get("user_login") for r in user_records
                    if r.get("user_login") and (r.get("date") or r.get("day", "")) == latest_date
                })
                engaged = len({r.get("user_login") for r in user_records if r.get("user_login")})
                sugg, acc = 0, 0
                for rec in user_records:
                    for feat in rec.get("totals_by_feature", []):
                        sugg += feat.get("code_generation_activity_count", 0)
                        acc += feat.get("code_acceptance_activity_count", 0)
                rate = f"{acc / sugg * 100:.1f}%" if sugg else "N/A"
                seats_label = "N/A"
                seat_result = await orch._tool_seat_info({"org": config.github_org})
                total = seat_result.get("seat_info", {}).get("total_seats", 0)
                if total:
                    seats_label = str(total)
            except Exception as e:
                logger.error("API metrics error: %s", e)
                return HTMLResponse(
                    f'<div class="kpi-card"><div class="kpi-value">!</div>'
                    f'<div class="kpi-label">{e}</div></div>'
                )

        html = (
            f'<div class="kpi-card">'
            f'  <div class="kpi-value">{active:,}</div>'
            f'  <div class="kpi-label">{t.get("dash_active_users", "Active Users")}</div>'
            f'</div>'
            f'<div class="kpi-card">'
            f'  <div class="kpi-value">{engaged:,}</div>'
            f'  <div class="kpi-label">{t.get("dash_engaged_users", "Engaged Users")}</div>'
            f'</div>'
            f'<div class="kpi-card">'
            f'  <div class="kpi-value">{rate}</div>'
            f'  <div class="kpi-label">{t.get("dash_acceptance_rate", "Acceptance Rate")}</div>'
            f'</div>'
            f'<div class="kpi-card">'
            f'  <div class="kpi-value">{seats_label}</div>'
            f'  <div class="kpi-label">{t.get("dash_active_seats", "Seats")}</div>'
            f'</div>'
        )
        return HTMLResponse(html)

    @app.get("/api/charts/adoption")
    async def api_chart_adoption(request: Request):
        """Return adoption-trend data for the Plotly line chart."""
        filter_logins = _resolve_filter_logins(request)
        if filter_logins is not None:
            # Filtered: aggregate from user-level data
            try:
                logins = filter_logins
                user_records = await _get_raw_user_metrics()
                filtered = [r for r in user_records if r.get("user_login", "").lower() in logins]
                by_day: dict[str, set[str]] = {}
                for r in filtered:
                    # After _unwrap_day_totals the date field is "date"; fall back to "day"
                    day = r.get("date") or r.get("day", "")
                    login = r.get("user_login", "")
                    if not day or not login:
                        continue
                    by_day.setdefault(day, set()).add(login)
                days_sorted = sorted(by_day.keys())
                # Active  = unique users on each individual day
                # Engaged = cumulative unique users from the start of the window up to
                #           each day (mirrors total_engaged_users rolling 28-day semantics)
                seen: set[str] = set()
                engaged_counts: list[int] = []
                for d in days_sorted:
                    seen.update(by_day[d])
                    engaged_counts.append(len(seen))
                return {
                    "dates": days_sorted,
                    "active_users": [len(by_day[d]) for d in days_sorted],
                    "engaged_users": engaged_counts,
                }
            except Exception as e:
                logger.error("Chart adoption (filtered) error: %s", e)
                return {"error": str(e)}

        orch = app.state.orchestrator
        if not orch:
            return {"error": "Orchestrator not initialized"}
        try:
            result = await orch._tool_org_metrics({
                "org": config.github_org,
                "period": "28-day",
            })
            metrics_list = result.get("metrics", [])
            # Sort by date ascending for the chart.
            metrics_list.sort(key=lambda m: m.get("date", ""))
            dates = [m.get("date", "") for m in metrics_list]
            active = [m.get("total_active_users", 0) for m in metrics_list]
            engaged = [m.get("total_engaged_users", 0) for m in metrics_list]
            return {"dates": dates, "active_users": active, "engaged_users": engaged}
        except Exception as e:
            logger.error("Chart adoption error: %s", e)
            return {"error": str(e)}

    @app.get("/api/charts/features")
    async def api_chart_features(request: Request):
        """Return feature adoption as unique-user counts (28 days).

        Uses user-level records so the metric is consistent regardless of filter:
        each bar shows how many distinct users touched that feature at least once.
        """
        _LABELS = {
            "code_completion": "Code Completions",
            "chat_panel_agent_mode": "Agent Mode",
            "agent_edit": "Agent Edits",
            "chat_panel_ask_mode": "Chat — Ask",
            "chat_panel_custom_mode": "Chat — Custom",
            "chat_panel_edit_mode": "Chat — Edit",
            "chat_panel_plan_mode": "Chat — Plan",
            "chat_panel_unknown_mode": "Chat — Other",
            "chat_inline": "Inline Chat",
            "copilot_cli": "Copilot CLI (feature)",
        }

        try:
            user_records = await _get_raw_user_metrics()
            if not user_records:
                return {"labels": [], "values": []}

            fl = _resolve_filter_logins(request)
            if fl is not None:
                records = [r for r in user_records if r.get("user_login", "").lower() in fl]
            else:
                records = user_records

            if not records:
                return {"labels": [], "values": []}

            # Unique users per feature — avoids scale mismatch between
            # completion suggestions (thousands) and chat interactions (tens).
            feature_users: dict[str, set[str]] = {}
            for rec in records:
                login = rec.get("user_login", "")
                if not login:
                    continue
                for feat in rec.get("totals_by_feature", []):
                    fname = feat.get("feature", "")
                    if not fname:
                        continue
                    has_activity = (
                        feat.get("code_generation_activity_count", 0) > 0
                        or feat.get("code_acceptance_activity_count", 0) > 0
                        or feat.get("user_initiated_interaction_count", 0) > 0
                    )
                    if has_activity:
                        label = _LABELS.get(fname, fname)
                        feature_users.setdefault(label, set()).add(login)
                # CLI sessions count as a feature
                cli_data = rec.get("totals_by_cli")
                if isinstance(cli_data, dict) and cli_data.get("session_count", 0) > 0:
                    feature_users.setdefault("CLI Sessions", set()).add(login)

            totals = {label: len(users) for label, users in feature_users.items()}

            sorted_items = sorted(
                ((l, v) for l, v in totals.items() if v > 0),
                key=lambda x: x[1],
                reverse=True,
            )
            if not sorted_items:
                return {"labels": [], "values": []}
            labels, values = zip(*sorted_items)
            return {"labels": list(labels), "values": list(values), "metric": "users"}
        except Exception as e:
            logger.error("Chart features error: %s", e)
            return {"error": str(e)}

    @app.get("/api/charts/top-users")
    async def api_chart_top_users(request: Request):
        """Return top 10 active users for a horizontal bar chart."""
        filter_logins = _resolve_filter_logins(request)

        try:
            user_records = await _get_raw_user_metrics()
            if not user_records:
                return {"logins": [], "scores": []}

            agg: dict[str, int] = {}
            for rec in user_records:
                login = rec.get("user_login", "")
                if not login:
                    continue
                if filter_logins is not None and login.lower() not in filter_logins:
                    continue
                for feat in rec.get("totals_by_feature", []):
                    fname = feat.get("feature", "")
                    if fname == "code_completion":
                        agg[login] = agg.get(login, 0) + feat.get("code_generation_activity_count", 0)
                    else:
                        count = feat.get("user_initiated_interaction_count", 0)
                        if count == 0:
                            count = feat.get("code_generation_activity_count", 0)
                        agg[login] = agg.get(login, 0) + count

            top = sorted(agg.items(), key=lambda x: x[1], reverse=True)[:10]
            top.reverse()
            return {
                "logins": [t[0] for t in top],
                "scores": [t[1] for t in top],
            }
        except Exception as e:
            logger.error("Chart top-users error: %s", e)
            return {"error": str(e)}

    @app.get("/api/charts/suggested-accepted")
    async def api_chart_suggested_accepted(request: Request):
        """Return daily suggested vs accepted code lines for the last 14 days."""
        try:
            records, is_filtered = await _get_filtered_records(request)
            if not records:
                return {"dates": [], "suggested": [], "accepted": []}

            # Aggregate by day
            day_sugg: dict[str, int] = {}
            day_acc: dict[str, int] = {}
            for rec in records:
                day = rec.get("day", rec.get("date", ""))
                if not day:
                    continue
                for feat in rec.get("totals_by_feature", []):
                    if feat.get("feature") == "code_completion":
                        day_sugg[day] = day_sugg.get(day, 0) + feat.get("code_generation_activity_count", 0)
                        day_acc[day] = day_acc.get(day, 0) + feat.get("code_acceptance_activity_count", 0)
                if not is_filtered:
                    comp = rec.get("copilot_ide_code_completions") or {}
                    s = comp.get("total_code_suggestions", 0)
                    a = comp.get("total_code_acceptances", 0)
                    if s:
                        day_sugg[day] = s
                        day_acc[day] = a

            days_sorted = sorted(day_sugg.keys())[-14:]
            return {
                "dates": days_sorted,
                "suggested": [day_sugg.get(d, 0) for d in days_sorted],
                "accepted": [day_acc.get(d, 0) for d in days_sorted],
            }
        except Exception as e:
            logger.error("Chart suggested-accepted error: %s", e)
            return {"error": str(e)}

    @app.get("/api/charts/usage-trend")
    async def api_chart_usage_trend(request: Request):
        """Return 28-day composite usage score trend."""
        try:
            records, is_filtered = await _get_filtered_records(request)
            if not records:
                return {"dates": [], "scores": []}

            day_scores: dict[str, int] = {}
            for rec in records:
                day = rec.get("day", rec.get("date", ""))
                if not day:
                    continue
                score = 0
                for feat in rec.get("totals_by_feature", []):
                    fname = feat.get("feature", "")
                    if fname == "code_completion":
                        score += feat.get("code_generation_activity_count", 0)
                    else:
                        s = feat.get("user_initiated_interaction_count", 0)
                        if s == 0:
                            s = feat.get("code_generation_activity_count", 0)
                        score += s
                if not is_filtered:
                    # Use org-level aggregate fields if available
                    comp = rec.get("copilot_ide_code_completions") or {}
                    ide_chat = rec.get("copilot_ide_chat") or {}
                    dot_chat = rec.get("copilot_dotcom_chat") or {}
                    pr = rec.get("copilot_dotcom_pull_requests") or {}
                    cli = rec.get("copilot_cli") or {}
                    org_score = (
                        comp.get("total_code_suggestions", 0)
                        + ide_chat.get("total_chats", 0)
                        + dot_chat.get("total_chats", 0)
                        + pr.get("total_pr_summaries_created", 0)
                        + cli.get("total_chats", 0)
                    )
                    if org_score:
                        score = org_score
                day_scores[day] = day_scores.get(day, 0) + score

            days_sorted = sorted(day_scores.keys())
            return {
                "dates": days_sorted,
                "scores": [day_scores[d] for d in days_sorted],
            }
        except Exception as e:
            logger.error("Chart usage-trend error: %s", e)
            return {"error": str(e)}

    @app.get("/api/roi-data")
    async def api_roi_data(request: Request):
        """Return raw data for the client-side ROI calculator."""
        try:
            records, is_filtered = await _get_filtered_records(request)
            if not records:
                return {"agent_edits": 0, "total_seats": 0, "days": 0}

            agent_edits = _sum_agent_edits(records)

            if is_filtered:
                active_users = _resolve_active_user_count(records, is_filtered=True)
                total_seats = await _resolve_total_licenses(request, is_filtered=True)
                unique_days = {r.get("day") for r in records if r.get("day")}
                days = len(unique_days)
            else:
                active_users = _resolve_active_user_count(records, is_filtered=False)
                days = len(records)
                total_seats = 0
                orch = app.state.orchestrator
                if orch:
                    try:
                        seat_result = await orch._tool_seat_info({"org": config.github_org})
                        si = seat_result.get("seat_info", {})
                        total_seats = si.get("total_seats", 0)
                    except Exception:
                        pass

            return {
                "agent_edits": agent_edits,
                "total_seats": total_seats,
                "active_users": active_users,
                "days": days,
            }
        except Exception as e:
            logger.error("ROI data error: %s", e)
            return {"error": str(e)}

    @app.get("/api/productivity-insights", response_class=HTMLResponse)
    async def api_productivity_insights(request: Request):
        """Return Productivity Insights KPI cards as an HTML fragment."""
        lang = _lang(request)
        t = get_translations(lang)

        def _fmt_ratio(value: float | None) -> str:
            return f"{value:.1f}" if value is not None else "N/A"

        try:
            records, is_filtered = await _get_filtered_records(request)
            agent_edits = _sum_agent_edits(records)
            active_users = _resolve_active_user_count(records, is_filtered=is_filtered)
            total_licenses = await _resolve_total_licenses(request, is_filtered=is_filtered)
            metrics = _compute_productivity_metrics(
                agent_edits=agent_edits,
                active_users=active_users,
                total_licenses=total_licenses,
            )
        except Exception as e:
            logger.error("Productivity insights error: %s", e)
            return HTMLResponse(
                f'<div class="productivity-card">'
                f'  <div class="productivity-label">{t.get("dash_productivity_title", "Productivity Insights")}</div>'
                f'  <div class="productivity-desc">{e}</div>'
                f'</div>'
            )

        active_users_base = (
            f'{metrics["active_users"]:,} {t.get("product_active_users", "Active Users")}'
            if metrics["active_users"]
            else "N/A"
        )
        total_licenses_base = (
            f'{metrics["total_licenses"]:,} {t.get("product_total_licenses", "Total Licenses")}'
            if metrics["total_licenses"]
            else "N/A"
        )
        badge_html = ""
        if metrics["efficacy_badge"]:
            badge_html = (
                f'<span class="productivity-badge">'
                f'{t.get("product_badge_advanced", "Advanced")}'
                f"</span>"
            )

        # Compute 28-day scaled thresholds for the active-users card.
        active_rate = _compute_active_rate(
            [metrics["active_users"]] if metrics["active_users"] else [],
            metrics["total_licenses"],
        )
        efficacy_thresholds = [round(th / active_rate, 1) for th in _CANONICAL_THRESHOLDS_28D]
        adoption_thresholds = list(_CANONICAL_THRESHOLDS_28D)

        efficacy_legend = _build_scale_legend_html(efficacy_thresholds, t)
        adoption_legend = _build_scale_legend_html(adoption_thresholds, t)

        html = (
            f'<div class="productivity-card">'
            f'  <div class="productivity-card-head">'
            f'    <div class="productivity-label">{t.get("product_efficiency_title", "Agent Effectiveness")}</div>'
            f'    {badge_html}'
            f'  </div>'
            f'  <div class="productivity-value">{_fmt_ratio(metrics["agent_efficacy"])}</div>'
            f'  <div class="productivity-formula">{metrics["agent_edits"]:,} {t.get("product_agent_edits", "Agent Edits")} / {active_users_base}</div>'
            f'  <div class="productivity-desc">{t.get("product_efficiency_desc", "Agent Edits divided by active users over the last 28 days.")}</div>'
            f'  {efficacy_legend}'
            f'</div>'
            f'<div class="productivity-card productivity-card-secondary">'
            f'  <div class="productivity-card-head">'
            f'    <div class="productivity-label">{t.get("product_real_adoption_title", "Real Adoption")}</div>'
            f'  </div>'
            f'  <div class="productivity-value">{_fmt_ratio(metrics["real_adoption"])}</div>'
            f'  <div class="productivity-formula">{metrics["agent_edits"]:,} {t.get("product_agent_edits", "Agent Edits")} / {total_licenses_base}</div>'
            f'  <div class="productivity-desc">{t.get("product_real_adoption_desc", "Agent Edits divided by total licenses to show breadth of adoption.")}</div>'
            f'  {adoption_legend}'
            f'</div>'
        )
        return HTMLResponse(html)

    @app.get("/api/charts/productivity-trend")
    async def api_chart_productivity_trend(request: Request):
        """Return weekly Agent Edits / User values for the last 13 weeks."""
        try:
            return await _get_weekly_agent_ratio_payload(request)
        except Exception as e:
            logger.error("Chart productivity-trend error: %s", e)
            return {"error": str(e)}

    @app.get("/api/charts/productivity-trend-seats")
    async def api_chart_productivity_trend_seats(request: Request):
        """Return weekly Agent Edits / Total Seats values for the last 13 weeks."""
        try:
            series = await _get_weekly_agent_series(request)

            # Resolve total seats — must respect the active org filter.
            fl_seats = _resolve_filter_logins(request)
            is_filtered = fl_seats is not None
            if is_filtered:
                # Filtered: denominator = employees in the selected org group.
                total_seats = len(fl_seats)
            else:
                # Unfiltered: use the org-wide seat count from the GitHub API.
                total_seats = 0
                orch = app.state.orchestrator
                if orch:
                    try:
                        seat_result = await orch._tool_seat_info({"org": config.github_org})
                        total_seats = seat_result.get("seat_info", {}).get("total_seats", 0)
                    except Exception:
                        pass
                if not total_seats:
                    total_seats = await _resolve_total_licenses(request, is_filtered=False)

            values: list[float] = []
            for edits in series["agent_edits"]:
                values.append(round(edits / total_seats, 1) if total_seats else 0.0)

            average = round(sum(values) / len(values), 1) if values else 0.0
            return {
                "labels": series["labels"],
                "values": values,
                "colors": [_agent_edit_weekly_color(v) for v in values],
                "average": average,
                "total_seats": total_seats,
                "thresholds": list(_CANONICAL_THRESHOLDS),
            }
        except Exception as e:
            logger.error("Chart productivity-trend-seats error: %s", e)
            return {"error": str(e)}

    @app.get("/api/charts/agent-edits-wow")
    async def api_chart_agent_edits_wow(request: Request):
        """Return weekly Agent Edits / User for the last 13 weeks."""
        try:
            payload = await _get_weekly_agent_ratio_payload(request)
            return {
                "labels": payload["labels"],
                "values": payload["values"],
                "colors": payload["colors"],
            }
        except Exception as e:
            logger.error("Chart agent-edits-wow error: %s", e)
            return {"error": str(e)}

    @app.get("/api/virtual-fte")
    async def api_virtual_fte(
        request: Request,
        lines_per_day: int = 50,
        working_days: int = 20,
        review_overhead: float = 0.20,
        hourly_rate: float = 33.0,
        daily_hours: int = 8,
        lines_per_agent_edit: int = 17,
    ):
        """Return Virtual FTE analysis combining IDE completions + Agent edits (last 13 weeks).

        Formula:
          volume_ide    = code_lines_accepted  (from IDE completions)
          volume_agent  = agent_edits × lines_per_agent_edit
          total_lines   = (volume_ide + volume_agent) × (1 − review_overhead)
          fte           = total_lines / (lines_per_day × working_days)
        """
        try:
            series = await _get_weekly_agent_series(request)
            weeks = _weekly_reference_dates()
            lines_series  = series.get("lines_accepted", [0] * len(weeks))
            edits_series  = series.get("agent_edits",   [0] * len(weeks))

            # Group weekly samples by calendar month.
            # Each Wednesday sample × 7 approximates the full-week contribution.
            monthly: dict[str, dict[str, Any]] = {}
            for week_date, lines_day, edits_day in zip(weeks, lines_series, edits_series):
                month_key = week_date.strftime("%b %Y")
                sort_key  = week_date.strftime("%Y-%m")
                if month_key not in monthly:
                    monthly[month_key] = {
                        "volume_ide": 0, "volume_agent": 0, "sort_key": sort_key
                    }
                monthly[month_key]["volume_ide"]   += lines_day  * 7
                monthly[month_key]["volume_agent"] += edits_day  * 7 * lines_per_agent_edit

            human_capacity   = lines_per_day * working_days
            monthly_dev_cost = hourly_rate * daily_hours * working_days

            periods = []
            for month_key in sorted(monthly, key=lambda k: monthly[k]["sort_key"]):
                d = monthly[month_key]
                vol_ide   = d["volume_ide"]
                vol_agent = d["volume_agent"]
                total_lines = (vol_ide + vol_agent) * (1.0 - review_overhead)
                fte   = round(total_lines / human_capacity, 2) if human_capacity else 0.0
                value = round(fte * monthly_dev_cost, 0)
                periods.append({
                    "period":       month_key,
                    "volume_ide":   int(vol_ide),
                    "volume_agent": int(vol_agent),
                    "total_lines":  int(total_lines),
                    "human_capacity": int(human_capacity),
                    "fte":          fte,
                    "monthly_value": int(value),
                })

            if not periods:
                return {
                    "periods": [],
                    "avg_fte": 0,
                    "monthly_dev_cost": int(monthly_dev_cost),
                    "human_capacity": int(human_capacity),
                }

            avg_fte = round(sum(p["fte"] for p in periods) / len(periods), 2)
            level, value = _parse_filter(request)
            total_seats = await _resolve_total_licenses(request, is_filtered=bool(level and value))

            return {
                "periods": periods,
                "avg_fte": avg_fte,
                "monthly_dev_cost": int(monthly_dev_cost),
                "human_capacity": int(human_capacity),
                "total_seats": total_seats,
            }
        except Exception as e:
            logger.error("Virtual FTE error: %s", e)
            return {"error": str(e)}

    @app.get("/api/adoption-kpis", response_class=HTMLResponse)
    async def api_adoption_kpis(request: Request):
        """Return CLI & Agent adoption KPIs as an HTML fragment."""
        lang = _lang(request)
        t = get_translations(lang)
        title = t.get("dash_adoption_title", "CLI & Agent Adoption")
        try:
            records, _ = await _get_filtered_records(request)
        except Exception as e:
            logger.error("Adoption KPIs error: %s", e)
            return HTMLResponse(
                f'<h3>{title}</h3>'
                f'<p style="color:var(--text-secondary)">{e}</p>'
            )

        if not records:
            return HTMLResponse(
                f'<h3>{title}</h3>'
                f'<p style="color:var(--text-secondary)">No data</p>'
            )

        sorted_recs = sorted(records, key=lambda r: r.get("day", ""))
        latest = sorted_recs[-1]
        total_users = latest.get("monthly_active_users", 0)
        # For filtered (user-level) data, compute counts from unique logins
        if not total_users:
            total_users = len({r.get("user_login") for r in sorted_recs if r.get("user_login")})

        # 1. CLI Users
        cli_users = latest.get("daily_active_cli_users", 0)
        if not cli_users:
            cli_users = len({
                r.get("user_login") for r in sorted_recs
                if r.get("user_login") and isinstance(r.get("totals_by_cli"), dict)
                and r["totals_by_cli"].get("session_count", 0) > 0
            })
        cli_users_pct = f"({cli_users / total_users * 100:.0f}%)" if total_users else ""

        # 2. CLI Sessions (28-day total)
        cli_sessions = 0
        for rec in sorted_recs:
            c = rec.get("totals_by_cli")
            if isinstance(c, dict):
                cli_sessions += c.get("session_count", 0)

        # 3. Agent Users
        agent_users = latest.get("monthly_active_agent_users", 0)
        if not agent_users:
            agent_users = len({
                r.get("user_login") for r in sorted_recs
                if r.get("user_login") and any(
                    f.get("feature") == "agent_edit" and f.get("code_generation_activity_count", 0) > 0
                    for f in r.get("totals_by_feature", [])
                )
            })
        agent_users_pct = f"({agent_users / total_users * 100:.0f}%)" if total_users else ""

        # 4. CLI Total Tokens (28-day)
        total_prompt_tokens = 0
        total_output_tokens = 0
        for rec in sorted_recs:
            c = rec.get("totals_by_cli")
            if isinstance(c, dict):
                tu = c.get("token_usage", {})
                total_prompt_tokens += tu.get("prompt_tokens_sum", 0)
                total_output_tokens += tu.get("output_tokens_sum", 0)
        total_tokens = total_prompt_tokens + total_output_tokens

        def _fmt(n: int) -> str:
            if n >= 1_000_000_000:
                return f"{n / 1_000_000_000:.1f}B"
            if n >= 1_000_000:
                return f"{n / 1_000_000:.1f}M"
            if n >= 1_000:
                return f"{n / 1_000:.1f}K"
            return str(n)

        html = (
            f'<h3>{title}</h3>'
            f'<div class="adoption-grid">'
            f'  <div class="adoption-card">'
            f'    <div class="adoption-icon">⌨</div>'
            f'    <div class="adoption-body">'
            f'      <div class="adoption-value">{cli_users:,} <span class="adoption-pct">{cli_users_pct}</span></div>'
            f'      <div class="adoption-label">{t.get("adopt_cli_users", "CLI Users")} <span class="adoption-period">({latest.get("day", "")})</span></div>'
            f'      <div class="adoption-desc">{t.get("adopt_cli_users_desc", "Developers who used gh copilot today")}</div>'
            f'    </div>'
            f'  </div>'
            f'  <div class="adoption-card">'
            f'    <div class="adoption-icon">▶</div>'
            f'    <div class="adoption-body">'
            f'      <div class="adoption-value">{cli_sessions:,}</div>'
            f'      <div class="adoption-label">{t.get("adopt_cli_sessions", "CLI Sessions")} <span class="adoption-period">(28d)</span></div>'
            f'      <div class="adoption-desc">{t.get("adopt_cli_sessions_desc", "Terminal sessions started with gh copilot")}</div>'
            f'    </div>'
            f'  </div>'
            f'  <div class="adoption-card">'
            f'    <div class="adoption-icon">🤖</div>'
            f'    <div class="adoption-body">'
            f'      <div class="adoption-value">{agent_users:,} <span class="adoption-pct">{agent_users_pct}</span></div>'
            f'      <div class="adoption-label">{t.get("adopt_agent_users", "Coding Agent Users")} <span class="adoption-period">(28d)</span></div>'
            f'      <div class="adoption-desc">{t.get("adopt_agent_users_desc", "Developers who used Copilot Agent for autonomous coding")}</div>'
            f'    </div>'
            f'  </div>'
            f'  <div class="adoption-card">'
            f'    <div class="adoption-icon">⚡</div>'
            f'    <div class="adoption-body">'
            f'      <div class="adoption-value">{_fmt(total_tokens)}</div>'
            f'      <div class="adoption-label">{t.get("adopt_cli_tokens", "CLI Tokens")} <span class="adoption-period">(28d)</span></div>'
            f'      <div class="adoption-desc">{t.get("adopt_cli_tokens_desc", "Total tokens exchanged in CLI sessions — high volume indicates complex tasks")}</div>'
            f'    </div>'
            f'  </div>'
            f'</div>'
        )
        return HTMLResponse(html)

    @app.get("/api/insights", response_class=HTMLResponse)
    async def api_insights(request: Request):
        """Return Quick Insights as an HTML fragment for HTMX swap."""
        lang = _lang(request)
        t = get_translations(lang)
        try:
            records, _is_filtered_ins = await _get_filtered_records(request)
        except Exception as e:
            logger.error("Insights error: %s", e)
            return HTMLResponse(
                f'<h3>{t.get("dash_insights_title", "Quick Insights")}</h3>'
                f'<p style="color:var(--text-secondary)">{e}</p>'
            )

        if not records:
            return HTMLResponse(
                f'<h3>{t.get("dash_insights_title", "Quick Insights")}</h3>'
                f'<p style="color:var(--text-secondary)">No data</p>'
            )

        # --- Top 5 languages (exclude "Other"/"Others"/"Unknown") ---
        _EXCLUDE_LANGS = {"other", "others", "unknown"}
        lang_counts: dict[str, int] = {}
        for rec in records:
            for lf in rec.get("totals_by_language_feature", []):
                name = lf.get("language", "")
                if name and name.lower() not in _EXCLUDE_LANGS:
                    lang_counts[name] = lang_counts.get(name, 0) + lf.get("code_generation_activity_count", 0)
        top5_langs = sorted(lang_counts.items(), key=lambda x: x[1], reverse=True)[:5]

        # --- Top 5 models ---
        model_counts: dict[str, int] = {}
        for rec in records:
            for mf in rec.get("totals_by_model_feature", []):
                name = mf.get("model", "")
                if name:
                    model_counts[name] = model_counts.get(name, 0) + mf.get("code_generation_activity_count", 0)
        top5_models = sorted(model_counts.items(), key=lambda x: x[1], reverse=True)[:5]

        # --- Top IDE (unique users per IDE) ---
        ide_users: dict[str, set[str]] = {}
        for rec in records:
            login = rec.get("user_login", "")
            for ide in rec.get("totals_by_ide", []):
                name = ide.get("ide", "")
                if name:
                    ide_users.setdefault(name, set()).add(login or f"_rec_{id(rec)}")
        top_ide = max(ide_users, key=lambda k: len(ide_users[k])) if ide_users else "N/A"

        # --- Acceptance rate (last 7 vs prev 7) ---
        sorted_recs = sorted(records, key=lambda r: r.get("day", ""))
        last7 = sorted_recs[-7:]
        prev7 = sorted_recs[-14:-7] if len(sorted_recs) >= 14 else []

        def _acc_rate(recs: list) -> float:
            sugg, acc = 0, 0
            for r in recs:
                for f in r.get("totals_by_feature", []):
                    if f.get("feature") == "code_completion":
                        sugg += f.get("code_generation_activity_count", 0)
                        acc += f.get("code_acceptance_activity_count", 0)
            return (acc / sugg * 100) if sugg else 0.0

        rate_last = _acc_rate(last7)
        rate_prev = _acc_rate(prev7)
        rate_delta = rate_last - rate_prev
        rate_arrow = "↑" if rate_delta > 0 else ("↓" if rate_delta < 0 else "→")
        rate_color = "var(--accent-green)" if rate_delta >= 0 else "#f85149"

        # --- Most active day ---
        day_activity: dict[str, int] = {}
        day_users: dict[str, set[str]] = {}
        for rec in sorted_recs:
            day = rec.get("day", "")
            dau = rec.get("daily_active_users", 0)
            if dau:
                day_activity[day] = dau
            else:
                login = rec.get("user_login", "")
                if login:
                    day_users.setdefault(day, set()).add(login)
        if not day_activity and day_users:
            day_activity = {d: len(u) for d, u in day_users.items()}
        best_day = max(day_activity, key=day_activity.get) if day_activity else "N/A"
        best_day_users = day_activity.get(best_day, 0)

        # Build top-5 languages HTML
        if top5_langs:
            total_lang = sum(v for _, v in top5_langs)
            lang_rows = "".join(
                f'<div class="lang-row">'
                f'<span class="lang-name">{name}</span>'
                f'<span class="lang-pct">{count / total_lang * 100:.0f}%</span>'
                f'</div>'
                for name, count in top5_langs
            )
            lang_card = (
                f'<div class="insight-card insight-card-tall">'
                f'  <span class="insight-label">{t.get("insight_top_langs", "Top 5 languages")}</span>'
                f'  <div class="lang-list">{lang_rows}</div>'
                f'</div>'
            )
        else:
            lang_card = (
                f'<div class="insight-card">'
                f'  <span class="insight-value">N/A</span>'
                f'  <span class="insight-label">{t.get("insight_top_langs", "Top 5 languages")}</span>'
                f'</div>'
            )

        # Build top-5 models HTML
        if top5_models:
            total_model = sum(v for _, v in top5_models)
            model_rows = "".join(
                f'<div class="lang-row">'
                f'<span class="lang-name">{name}</span>'
                f'<span class="lang-pct">{count / total_model * 100:.0f}%</span>'
                f'</div>'
                for name, count in top5_models
            )
            model_card = (
                f'<div class="insight-card insight-card-tall">'
                f'  <span class="insight-label">{t.get("insight_top_models", "Top 5 models")}</span>'
                f'  <div class="lang-list">{model_rows}</div>'
                f'</div>'
            )
        else:
            model_card = (
                f'<div class="insight-card">'
                f'  <span class="insight-value">N/A</span>'
                f'  <span class="insight-label">{t.get("insight_top_models", "Top 5 models")}</span>'
                f'</div>'
            )

        html = (
            f'<h3>{t.get("dash_insights_title", "Quick Insights")}</h3>'
            f'<div class="insights-grid">'
            f'  {lang_card}'
            f'  {model_card}'
            f'  <div class="insight-card">'
            f'    <span class="insight-value">{top_ide}</span>'
            f'    <span class="insight-label">{t.get("insight_top_ide", "Top IDE")}</span>'
            f'  </div>'
            f'  <div class="insight-card">'
            f'    <span class="insight-value">{rate_last:.1f}% <span style="color:{rate_color};font-size:14px">{rate_arrow} {abs(rate_delta):.1f}pp</span></span>'
            f'    <span class="insight-label">{t.get("insight_acc_trend", "Acceptance rate (7d trend)")}</span>'
            f'  </div>'
            f'  <div class="insight-card">'
            f'    <span class="insight-value">{best_day_users}</span>'
            f'    <span class="insight-label">{t.get("insight_peak_day", "Peak day")} ({best_day})</span>'
            f'  </div>'
            f'</div>'
        )
        return HTMLResponse(html)

    @app.get("/api/metrics/json")
    async def api_metrics_json():
        """Raw JSON metrics endpoint (for programmatic access)."""
        orch = app.state.orchestrator
        if not orch:
            return {"error": "Orchestrator not initialized"}

        try:
            result = await orch._tool_org_metrics({
                "org": config.github_org,
                "period": "28-day",
            })
            return result
        except Exception as e:
            logger.error("API metrics error: %s", e)
            return {"error": str(e)}

    @app.get("/api/seat-info")
    async def api_seat_info():
        """API endpoint for seat information."""
        orch = app.state.orchestrator
        if not orch:
            return {"error": "Orchestrator not initialized"}

        try:
            result = await orch._tool_seat_info({"org": config.github_org})
            return result
        except Exception as e:
            logger.error("API seat info error: %s", e)
            return {"error": str(e)}

    @app.get("/api/inactive-users/xlsx")
    async def api_inactive_users_csv(request: Request, days: int = 14):
        """Download an XLSX of seat holders inactive in the past N days (default 14).

        Joins three sources:
        - Seat API  — full seat list (catches zero-activity users)
        - 28-day user metrics (cached) — last-seen date + activity stats
        - OrgDatabase — full org structure per user (github_id lookup)

        Respects the active org filter (level/value query params).
        """
        from datetime import date as _date, timedelta as _td, datetime as _dt
        days = max(1, min(days, 90))  # clamp to sensible range

        orch = app.state.orchestrator
        if not orch:
            return JSONResponse({"error": "Orchestrator not initialized"}, status_code=503)

        try:
            # ── 1. Seat list ──────────────────────────────────────────
            seat_result = await orch._tool_seat_info({"org": config.github_org})
            seats: list[dict[str, Any]] = seat_result.get("seat_info", {}).get("seats", [])

            # ── 2. 28-day user metrics (cached) ───────────────────────
            user_records = await _get_raw_user_metrics()
            org_records = await _get_raw_org_metrics()
            latest_date_str = max(
                (r.get("date") or r.get("day", "") for r in org_records), default=""
            )
            if latest_date_str:
                latest_date = _date.fromisoformat(latest_date_str)
                cutoff = latest_date - _td(days=days - 1)   # inclusive window
            else:
                cutoff = _date.today() - _td(days=days)

            # logins active in last N days
            active_last_14: set[str] = set()
            last_seen: dict[str, str] = {}
            suggestions_28: dict[str, int] = {}
            acceptances_28: dict[str, int] = {}

            for rec in user_records:
                login = (rec.get("user_login") or "").lower()
                if not login:
                    continue
                day_str = rec.get("date") or rec.get("day", "")
                if day_str:
                    if day_str > last_seen.get(login, ""):
                        last_seen[login] = day_str
                    try:
                        if _date.fromisoformat(day_str) >= cutoff:
                            active_last_14.add(login)
                    except ValueError:
                        pass
                for feat in rec.get("totals_by_feature", []):
                    suggestions_28[login] = suggestions_28.get(login, 0) + feat.get("code_generation_activity_count", 0)
                    acceptances_28[login] = acceptances_28.get(login, 0) + feat.get("code_acceptance_activity_count", 0)

            # ── 3. Org filter ─────────────────────────────────────────
            filter_logins: set[str] | None = _resolve_filter_logins(request)

            # ── 4. Build inactive list ────────────────────────────────
            inactive_seats = [
                s for s in seats
                if (s.get("login") or "").lower() not in active_last_14
                and (filter_logins is None or (s.get("login") or "").lower() in filter_logins)
            ]

            # ── 5. OrgDB lookup ───────────────────────────────────────
            db = _get_orgdb()
            def _org_row(login: str) -> dict[str, str]:
                try:
                    row = db._conn.execute(
                        "SELECT name, surname, email, business_title, job_family, "
                        "location, location_country, is_manager, "
                        "sup_org_level_4, sup_org_level_5, sup_org_level_6, "
                        "sup_org_level_7, sup_org_level_8, "
                        "hr_business_partner, cost_center_id "
                        "FROM employees WHERE LOWER(github_id) = ?",
                        (login.lower(),)
                    ).fetchone()
                    if row:
                        keys = ["name", "surname", "email", "business_title", "job_family",
                                "location", "location_country", "is_manager",
                                "sup_org_level_4", "sup_org_level_5", "sup_org_level_6",
                                "sup_org_level_7", "sup_org_level_8",
                                "hr_business_partner", "cost_center_id"]
                        return {k: ("" if v is None else str(v)) for k, v in zip(keys, row)}
                except Exception:
                    pass
                return {}

            # ── 6. Build CSV ──────────────────────────────────────────
            headers = [
                "github_login", "last_activity_at", "last_activity_editor", "assigned_at",
                "last_seen_in_metrics", "suggestions_28d", "acceptances_28d", "acceptance_rate_28d",
                "name", "surname", "email", "business_title", "job_family",
                "location", "location_country", "is_manager",
                "sup_org_level_4", "sup_org_level_5", "sup_org_level_6",
                "sup_org_level_7", "sup_org_level_8",
                "hr_business_partner", "cost_center_id",
            ]

            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Inactive Users"

            # Header row — blue background, white bold text
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill("solid", fgColor="2563EB")
            for col_idx, h in enumerate(headers, start=1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Data rows
            for row_idx, seat in enumerate(
                sorted(inactive_seats, key=lambda s: (s.get("login") or "").lower()),
                start=2,
            ):
                login = (seat.get("login") or "").lower()
                sugg = suggestions_28.get(login, 0)
                acc = acceptances_28.get(login, 0)
                rate = f"{acc / sugg * 100:.1f}%" if sugg else ""
                org = _org_row(login)
                row_values = [
                    seat.get("login", ""),
                    seat.get("last_activity_at") or "",
                    seat.get("last_activity_editor") or "",
                    seat.get("assigned_at") or "",
                    last_seen.get(login, ""),
                    sugg,
                    acc,
                    rate,
                    org.get("name", ""),
                    org.get("surname", ""),
                    org.get("email", ""),
                    org.get("business_title", ""),
                    org.get("job_family", ""),
                    org.get("location", ""),
                    org.get("location_country", ""),
                    org.get("is_manager", ""),
                    org.get("sup_org_level_4", ""),
                    org.get("sup_org_level_5", ""),
                    org.get("sup_org_level_6", ""),
                    org.get("sup_org_level_7", ""),
                    org.get("sup_org_level_8", ""),
                    org.get("hr_business_partner", ""),
                    org.get("cost_center_id", ""),
                ]
                for col_idx, val in enumerate(row_values, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=val)

            # Auto-size columns
            for col in ws.columns:
                max_len = max((len(str(c.value or "")) for c in col), default=0)
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 40)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            filename = f"inactive-users-{days}d-{_date.today().isoformat()}.xlsx"
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )

        except Exception as e:
            logger.error("Inactive users XLSX error: %s", e)
            return JSONResponse({"error": str(e)}, status_code=500)

    # ------------------------------------------------------------------
    # Setup page — org structure import & GitHub ID mapping
    # ------------------------------------------------------------------

    def _get_orgdb() -> OrgDatabase:
        """Get or create the OrgDatabase instance."""
        if not hasattr(app.state, "orgdb") or app.state.orgdb is None:
            app.state.orgdb = OrgDatabase()
        return app.state.orgdb

    _reload_org_map()

    @app.get("/setup", response_class=HTMLResponse)
    async def setup_page(request: Request):
        """Setup page for org structure import and user mapping."""
        db = _get_orgdb()
        stats = db.mapping_stats()
        return templates.TemplateResponse(
            "setup.html",
            _ctx(request, title="Copilot Pulse — Setup", active_tab="setup", stats=stats),
        )

    # ── Settings page ─────────────────────────────────────────

    def _mask_key(key: str) -> str:
        """Return a masked version of an API key (last 4 chars visible)."""
        if not key or len(key) < 8:
            return key
        return "•" * (len(key) - 4) + key[-4:]

    @app.get("/settings", response_class=HTMLResponse)
    async def settings_page(request: Request):
        """LLM provider settings page."""
        overrides = load_settings_file()
        current_provider = overrides.get("llm_provider") or config.llm_provider
        current_model = overrides.get("llm_model") or config.llm_model
        current_endpoint = overrides.get("llm_endpoint") or config.llm_endpoint
        raw_key = overrides.get("anthropic_api_key") or config.anthropic_api_key
        raw_llm_token = overrides.get("llm_github_token") or ""
        return templates.TemplateResponse(
            "settings.html",
            _ctx(
                request,
                title="Copilot Pulse — Settings",
                active_tab="settings",
                current_provider=current_provider,
                current_model=current_model,
                current_endpoint=current_endpoint,
                anthropic_key_masked=_mask_key(raw_key),
                llm_github_token_masked=_mask_key(raw_llm_token),
            ),
        )

    @app.post("/api/settings")
    async def api_save_settings(request: Request):
        """Validate, persist, and hot-swap the LLM provider."""
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"ok": False, "error": "Invalid JSON body"}, status_code=400)

        provider = body.get("llm_provider", "").strip().lower()
        if provider not in ("anthropic", "github-copilot"):
            return JSONResponse({"ok": False, "error": f"Invalid provider: '{provider}'"}, status_code=400)

        def _is_masked(value: str) -> bool:
            """Return True if the value contains masking characters (not a real key)."""
            return bool(value) and "•" in value

        anthropic_api_key = body.get("anthropic_api_key", "").strip()
        llm_model = body.get("llm_model", "").strip()
        llm_endpoint = body.get("llm_endpoint", "").strip()
        llm_github_token = body.get("llm_github_token", "").strip()

        overrides = load_settings_file()

        # For Anthropic: use submitted key if non-empty and not masked, else fall back to current key.
        if provider == "anthropic":
            if not anthropic_api_key or _is_masked(anthropic_api_key):
                anthropic_api_key = overrides.get("anthropic_api_key") or config.anthropic_api_key
            if not anthropic_api_key:
                return JSONResponse(
                    {"ok": False, "error": "Anthropic API key is required for the Anthropic provider."},
                    status_code=400,
                )
        elif provider == "github-copilot":
            # Use submitted token, or fall back to the one stored in settings.json.
            # NOTE: config.github_token is intentionally NOT used here — it may be empty
            # when GitHub App auth is active, and mixing them would break API auth.
            if not llm_github_token or _is_masked(llm_github_token):
                llm_github_token = overrides.get("llm_github_token") or ""
            if not llm_github_token:
                return JSONResponse(
                    {
                        "ok": False,
                        "error": (
                            "A GitHub token with Copilot access is required for the GitHub Copilot LLM provider. "
                            "Enter a Personal Access Token (PAT) with 'copilot' scope. "
                            "This token is used only for LLM calls and does not affect GitHub API authentication."
                        ),
                    },
                    status_code=400,
                )

        # Validate by instantiating the provider (catches bad key format etc.).
        try:
            new_provider = create_provider_from_settings(
                llm_provider=provider,
                anthropic_api_key=anthropic_api_key,
                github_token=llm_github_token,  # LLM-only token, never touches GitHub API auth
                llm_model=llm_model,
                llm_endpoint=llm_endpoint,
            )
        except Exception as exc:
            return JSONResponse({"ok": False, "error": str(exc)}, status_code=400)

        # Persist to settings.json.
        save_settings_file({
            "llm_provider": provider,
            "anthropic_api_key": anthropic_api_key,
            "llm_model": llm_model,
            "llm_endpoint": llm_endpoint,
            "llm_github_token": llm_github_token,
        })

        # Hot-swap the orchestrator's provider (no restart required).
        orch = app.state.orchestrator
        if orch:
            orch._provider = new_provider
            logger.info("LLM provider hot-swapped to '%s' (model=%r)", provider, llm_model or "default")

        return JSONResponse({"ok": True, "provider": provider})

    @app.post("/api/preview-org")
    async def api_preview_org(file: UploadFile = File(...)):
        """Step 1 of the two-step import: upload file and return column preview.

        Saves the file to a temp location and returns:
        - session_id: use this in /api/import-org to avoid re-uploading
        - columns: list of column descriptors (header, field, status, required)
        - sample_rows: first 10 data rows
        - missing_required: required field names absent from the file
        """
        if not file.filename or not file.filename.endswith((".xlsx", ".xls", ".csv")):
            return JSONResponse({"error": "Il file deve essere in formato .xlsx, .xls o .csv"}, status_code=400)

        _cleanup_upload_sessions()

        orig_suffix = Path(file.filename).suffix.lower() if file.filename else ".xlsx"
        with tempfile.NamedTemporaryFile(suffix=orig_suffix, delete=False) as tmp:
            shutil.copyfileobj(file.file, tmp)
            tmp_path = tmp.name

        try:
            from ..orgdata.loader import parse_xls_preview
            preview = parse_xls_preview(tmp_path)
        except Exception as e:
            Path(tmp_path).unlink(missing_ok=True)
            logger.error("Preview org error: %s", e)
            return JSONResponse({"error": str(e)}, status_code=500)

        session_id = str(uuid.uuid4())
        _upload_sessions[session_id] = {"path": tmp_path, "created_at": time.time()}
        preview["session_id"] = session_id
        return JSONResponse(preview)

    @app.post("/api/import-org")
    async def api_import_org(request: Request):
        """Step 2 of the two-step import: import with selected columns.

        Expects JSON body:
            session_id: str — returned by /api/preview-org
            selected_columns: list[str] — header strings to import (others skipped)
            refresh_all: bool — if true, wipes existing data first
        """
        try:
            body = await request.json()
        except Exception:
            return JSONResponse({"error": "Invalid JSON body"}, status_code=400)

        session_id = body.get("session_id", "")
        selected_columns: list[str] = body.get("selected_columns") or []
        refresh_all: bool = bool(body.get("refresh_all", False))

        session = _upload_sessions.get(session_id)
        if not session:
            return JSONResponse(
                {"error": "Sessione scaduta o non trovata. Ricarica il file e riprova."},
                status_code=400,
            )

        tmp_path = session["path"]

        try:
            from ..orgdata.loader import OrgDataLoader

            loader = OrgDataLoader(tmp_path)
            employees = loader.load(selected_headers=selected_columns or None)

            db = _get_orgdb()
            dicts = [e.model_dump() for e in employees]

            if refresh_all:
                count = db.import_employees(dicts)
                result: dict[str, Any] = {"imported": count, "preserved_mappings": 0}
            else:
                result = db.merge_employees(dicts)

            orch = app.state.orchestrator
            if orch and not orch._orgdb:
                orch._orgdb = db

            # Refresh the in-memory org filter map from the updated DB
            _reload_org_map()

            stats = db.mapping_stats()
            return JSONResponse({
                "success": True,
                "imported": result.get("imported", 0),
                "preserved_mappings": result.get("preserved_mappings", 0),
                "refresh_all": refresh_all,
                "stats": stats,
            })
        except Exception as e:
            logger.error("Import org error: %s", e)
            return JSONResponse({"error": str(e)}, status_code=500)
        finally:
            # Clean up session and temp file
            _upload_sessions.pop(session_id, None)
            Path(tmp_path).unlink(missing_ok=True)

    @app.get("/api/mapping-stats")
    async def api_mapping_stats():
        """Get current mapping statistics."""
        db = _get_orgdb()
        return db.mapping_stats()

    @app.get("/api/unmatched-users")
    async def api_unmatched_users():
        """Get list of unmatched GitHub users."""
        db = _get_orgdb()
        unmatched = db.unmatched_github_users()
        return {"unmatched": unmatched, "count": len(unmatched)}

    @app.post("/api/map-users/auto")
    async def api_auto_map(request: Request):
        """Run auto-mapping of GitHub users to employees."""
        db = _get_orgdb()
        rows = db._conn.execute(
            "SELECT DISTINCT github_login FROM copilot_usage"
        ).fetchall()
        github_users = [r["github_login"] for r in rows]

        if not github_users:
            return JSONResponse(
                {"error": "No Copilot usage data found. Use the Chat first to fetch user data."},
                status_code=400,
            )

        # Read optional pattern config from request body
        try:
            body = await request.json()
        except Exception:
            body = {}
        email_pattern = body.get("email_pattern", "{name}.{surname}")
        duplicate_strategy = body.get("duplicate_strategy", "skip")

        # Validate inputs
        allowed_patterns = ["{name}.{surname}", "{surname}.{name}", "{name1}.{surname}"]
        if email_pattern not in allowed_patterns:
            email_pattern = "{name}.{surname}"
        if duplicate_strategy not in ("skip", "seq2", "first"):
            duplicate_strategy = "skip"

        matches = db.auto_map_by_email(
            github_users,
            email_pattern=email_pattern,
            duplicate_strategy=duplicate_strategy,
        )
        if matches:
            _reload_org_map()
        stats = db.mapping_stats()
        return {
            "success": True,
            "new_matches": len(matches),
            "matches": matches,
            "stats": stats,
        }

    @app.post("/api/map-users/manual")
    async def api_manual_map(
        employee_id: str = Form(...),
        github_login: str = Form(...),
    ):
        """Manually map an employee to a GitHub login."""
        db = _get_orgdb()

        # Verify employee exists
        row = db._conn.execute(
            "SELECT employee_id, name, surname FROM employees WHERE employee_id = ?",
            (employee_id,),
        ).fetchone()
        if not row:
            return JSONResponse(
                {"error": f"Employee ID '{employee_id}' non trovato"},
                status_code=404,
            )

        db.set_github_id(employee_id, github_login, method="manual")
        _reload_org_map()
        stats = db.mapping_stats()
        return {
            "success": True,
            "mapped": f"{row['name']} {row['surname']} → {github_login}",
            "stats": stats,
        }

    @app.get("/api/employees/search")
    async def api_search_employees(q: str = ""):
        """Search employees by name, surname, or ID (for autocomplete)."""
        if len(q) < 2:
            return {"results": []}
        db = _get_orgdb()
        rows = db._conn.execute(
            """SELECT employee_id, name, surname, email_work, github_id
               FROM employees
               WHERE LOWER(name) LIKE LOWER(?) OR LOWER(surname) LIKE LOWER(?)
                  OR LOWER(employee_id) LIKE LOWER(?)
               LIMIT 20""",
            (f"%{q}%", f"%{q}%", f"%{q}%"),
        ).fetchall()
        return {
            "results": [
                {
                    "employee_id": r["employee_id"],
                    "name": r["name"],
                    "surname": r["surname"],
                    "email": r["email_work"],
                    "github_id": r["github_id"],
                }
                for r in rows
            ]
        }

    @app.websocket("/ws/chat")
    async def websocket_chat(websocket: WebSocket):
        """WebSocket endpoint for real-time chat with the agent.

        Streams ``type="log"`` messages for each orchestrator step so the
        browser can show them in an expandable details section.
        """
        await websocket.accept()
        orch = app.state.orchestrator

        if not orch:
            await websocket.send_json({"type": "error", "message": "Agent not initialized"})
            await websocket.close()
            return

        async def _send_log(message: str) -> None:
            """Forward orchestrator log entries to the browser."""
            try:
                await websocket.send_json({"type": "log", "message": message})
            except Exception:
                pass  # Connection may have closed mid-processing.

        try:
            while True:
                data = await websocket.receive_text()
                message = json.loads(data)
                question = message.get("question", "")

                if not question:
                    continue

                await websocket.send_json({"type": "status", "message": ""})

                # Attach log callback so the orchestrator streams steps to the
                # browser and suppresses terminal output for this request.
                orch._log_callback = _send_log
                try:
                    response = await orch.ask(question)
                    await websocket.send_json({
                        "type": "response",
                        "message": response,
                    })
                except Exception as e:
                    logger.error("Chat error: %s", e)
                    await websocket.send_json({
                        "type": "error",
                        "message": f"Errore: {e}",
                    })
                finally:
                    orch._log_callback = None

        except WebSocketDisconnect:
            orch._log_callback = None
            logger.info("WebSocket client disconnected")

    # ------------------------------------------------------------------
    # AI Maturity Radar — classification + trend endpoints
    # ------------------------------------------------------------------

    def _classify_user(
        agent: int,
        cli: int,
        chat: int,
        completions: int,
        lines_accepted: int,
        active_days: int,
    ) -> int:
        """Assign a maturity level 1–5 based on combined agentic activity (agent_turns + cli_turns)."""
        agentic = agent + cli
        if agentic >= 151:  return 5   # > 150
        if agentic >= 51:   return 4   # 51–150
        if agentic >= 11:   return 3   # 11–50
        if agentic >= 1:    return 2   # 1–10
        return 1                        # 0

    _LEVEL_NAMES = {
        5: "L5 Elite Agentic",
        4: "L4 Advanced Pilot",
        3: "L3 Hybrid Explorer",
        2: "L2 Traditionalist",
        1: "L1 Passive User",
    }

    @app.get("/maturity", response_class=HTMLResponse)
    async def maturity_page(request: Request):
        """AI Maturity Radar page."""
        return templates.TemplateResponse(
            "ai-maturity-radar.html",
            _ctx(
                request,
                title="AI Maturity Radar — Copilot Pulse",
                active_tab="maturity",
                has_org_filters=bool(_org_map),
            ),
        )

    @app.get("/api/maturity/classification")
    async def api_maturity_classification(request: Request, filter: str = "licensed"):
        """Classify all users/seats into 5 maturity levels.

        Query params:
          filter  = "active" | "licensed"  (default: "licensed")
          filter_level, filter_value       — org filter
        """
        try:
            orch = app.state.orchestrator
            filter_logins: set[str] | None = _resolve_filter_logins(request)

            # 1. Seat list → licensed logins
            licensed: set[str] = set()
            if orch:
                seat_result = await orch._tool_seat_info({"org": config.github_org})
                seats = seat_result.get("seat_info", {}).get("seats", [])
                licensed = {(s.get("login") or "").lower() for s in seats if s.get("login")}

            # 2. Per-user 28-day records
            user_records = await _get_raw_user_metrics()

            # Aggregate per-user
            agg: dict[str, dict[str, Any]] = {}
            for rec in user_records:
                login = (rec.get("user_login") or rec.get("github_login") or rec.get("login") or "").lower()
                if not login:
                    continue
                if login not in agg:
                    agg[login] = {
                        "agent": 0, "cli": 0, "chat": 0,
                        "completions": 0, "lines_accepted": 0,
                        "dates": set(),
                    }
                u = agg[login]
                day = rec.get("date") or rec.get("day", "")
                if day:
                    u["dates"].add(day)

                # Agent turns — legacy field OR totals_by_feature (GA)
                agent_val = rec.get("agent_turns", 0) or 0
                # CLI turns — legacy field; totals_by_cli dict handled below
                cli_val = rec.get("cli_turns", 0) or 0
                # Chat turns — legacy field OR totals_by_feature (GA)
                chat_val = rec.get("chat_turns", 0) or rec.get("user_initiated_interaction_count", 0) or 0
                # Completions — legacy field OR totals_by_feature (GA)
                comp_val = rec.get("completions_acceptances", 0) or 0
                lines_val = rec.get("completions_lines_accepted", 0) or 0

                for feat in rec.get("totals_by_feature", []):
                    fname = feat.get("feature", "")
                    if fname == "agent_edit":
                        agent_val += feat.get("code_generation_activity_count", 0)
                    elif fname == "code_completion":
                        comp_val += feat.get("code_acceptance_activity_count", 0)
                        lines_val += feat.get("loc_added_sum", 0)
                    elif "chat" in fname:
                        chat_val += feat.get("user_initiated_interaction_count", 0)

                # Fallback to top-level GA fields (code_acceptance_activity_count)
                # if completions weren't captured via totals_by_feature
                if not comp_val:
                    comp_val = rec.get("code_acceptance_activity_count", 0) or 0
                    lines_val = lines_val or rec.get("loc_added_sum", 0) or 0

                # totals_by_cli is a dict (not a list)
                _cli_data = rec.get("totals_by_cli")
                if isinstance(_cli_data, dict):
                    cli_val += _cli_data.get("session_count", 0)

                u["agent"] += agent_val
                u["cli"] += cli_val
                u["chat"] += chat_val
                u["completions"] += comp_val
                u["lines_accepted"] += lines_val

            # 3. Determine universe (active = has user_records, licensed = all seats)
            if filter == "active":
                universe = set(agg.keys())
            else:
                # Licensed: union of seats and any active user (active users not in seat list are rare)
                universe = licensed | set(agg.keys())

            # Apply org filter
            if filter_logins is not None:
                universe = universe & filter_logins

            # 4. Classify
            distribution: dict[str, int] = {"L1": 0, "L2": 0, "L3": 0, "L4": 0, "L5": 0}
            leaderboard_rows: list[dict[str, Any]] = []

            for login in universe:
                u = agg.get(login, {})
                lvl = _classify_user(
                    agent=u.get("agent", 0),
                    cli=u.get("cli", 0),
                    chat=u.get("chat", 0),
                    completions=u.get("completions", 0),
                    lines_accepted=u.get("lines_accepted", 0),
                    active_days=len(u.get("dates", set())),
                )
                distribution[f"L{lvl}"] += 1
                leaderboard_rows.append({
                    "login": login,
                    "level": lvl,
                    "agent_turns": u.get("agent", 0),
                    "cli_turns": u.get("cli", 0),
                    "active_days": len(u.get("dates", set())),
                })

            total = sum(distribution.values())
            skill_gap = (distribution["L1"] + distribution["L2"]) / total if total else 0.0
            champion_density = distribution["L5"] / total if total else 0.0

            # Top 10 leaderboard sorted by agent_turns
            top10 = sorted(leaderboard_rows, key=lambda r: r["agent_turns"], reverse=True)[:10]
            for i, row in enumerate(top10, 1):
                row["rank"] = i

            result = {
                "distribution": distribution,
                "total": total,
                "skill_gap_index": round(skill_gap, 4),
                "champion_density": round(champion_density, 4),
                "leaderboard": top10,
            }
            logger.info(
                "Maturity classification computed (users=%d, filter=%s, active_filters=%s)",
                total, filter, bool(filter_logins),
            )
            return JSONResponse(result)

        except Exception as e:
            logger.error("Maturity classification error: %s", e, exc_info=True)
            return JSONResponse({"error": str(e)}, status_code=500)

    @app.get("/api/maturity/search")
    async def api_maturity_search(request: Request, q: str = "", filter: str = "licensed"):
        """Search for a user by name, surname, email, or GitHub login and return their tier.

        Returns up to 10 matches.  Tier is computed from the cached 28-day user records.
        """
        q = q.strip()
        if len(q) < 2:
            return JSONResponse({"results": []})
        try:
            q_lower = q.lower()

            # 1. Aggregate per-user metrics (in-memory cached)
            user_records = await _get_raw_user_metrics()
            agg: dict[str, dict[str, Any]] = {}
            for rec in user_records:
                login = (
                    rec.get("user_login") or rec.get("github_login") or
                    rec.get("login") or ""
                ).lower()
                if not login:
                    continue
                if login not in agg:
                    agg[login] = {
                        "agent": 0, "cli": 0, "chat": 0,
                        "completions": 0, "lines_accepted": 0, "dates": set(),
                    }
                u = agg[login]
                day = rec.get("date") or rec.get("day", "")
                if day:
                    u["dates"].add(day)
                agent_v = rec.get("agent_turns", 0) or 0
                cli_v   = rec.get("cli_turns",   0) or 0
                chat_v  = rec.get("chat_turns",  0) or 0
                comp_v  = rec.get("completions_acceptances", 0) or 0
                lines_v = rec.get("completions_lines_accepted", 0) or 0
                for feat in rec.get("totals_by_feature", []):
                    fn = feat.get("feature", "")
                    if fn == "agent_edit":
                        agent_v += feat.get("code_generation_activity_count", 0)
                    elif fn == "code_completion":
                        comp_v  += feat.get("code_acceptance_activity_count", 0)
                        lines_v += feat.get("loc_added_sum", 0)
                    elif "chat" in fn:
                        chat_v  += feat.get("user_initiated_interaction_count", 0)
                if not comp_v:
                    comp_v  = rec.get("code_acceptance_activity_count", 0) or 0
                    lines_v = lines_v or rec.get("loc_added_sum", 0) or 0
                _cli_data = rec.get("totals_by_cli")
                if isinstance(_cli_data, dict):
                    cli_v += _cli_data.get("session_count", 0)
                u["agent"] += agent_v;  u["cli"] += cli_v
                u["chat"]  += chat_v;   u["completions"] += comp_v
                u["lines_accepted"] += lines_v

            # 2. Search OrgDB by name, surname, email, github_id
            db = _get_orgdb()
            like = f"%{q_lower}%"
            try:
                org_rows = db._conn.execute(
                    "SELECT github_id, name, surname, email FROM employees "
                    "WHERE LOWER(name) LIKE ? OR LOWER(surname) LIKE ? "
                    "OR LOWER(email) LIKE ? OR LOWER(github_id) LIKE ? "
                    "OR LOWER(name || ' ' || surname) LIKE ? "
                    "OR LOWER(surname || ' ' || name) LIKE ? "
                    "LIMIT 20",
                    (like, like, like, like, like, like)
                ).fetchall()
            except Exception:
                org_rows = []

            # Build candidate set: from org search + direct github_login match
            candidates: dict[str, dict[str, str]] = {}
            for github_id, name, surname, email in org_rows:
                if github_id:
                    candidates[github_id.lower()] = {
                        "name": name or "", "surname": surname or "", "email": email or ""
                    }
            # Also match logins directly in agg (only useful for single-word queries)
            if " " not in q_lower:
                for login in agg:
                    if q_lower in login and login not in candidates:
                        candidates[login] = {"name": "", "surname": "", "email": ""}

            # 3. Classify and build results
            results: list[dict[str, Any]] = []
            for login, org in candidates.items():
                if len(results) >= 10:
                    break
                u = agg.get(login, {})
                lvl = _classify_user(
                    agent=u.get("agent", 0),
                    cli=u.get("cli", 0),
                    chat=u.get("chat", 0),
                    completions=u.get("completions", 0),
                    lines_accepted=u.get("lines_accepted", 0),
                    active_days=len(u.get("dates", set())),
                )
                display = " ".join(x for x in [org["name"], org["surname"]] if x) or login
                results.append({
                    "login":      login,
                    "name":       org["name"],
                    "surname":    org["surname"],
                    "email":      org["email"],
                    "display":    display,
                    "tier_level": lvl,
                    "tier_name":  _LEVEL_NAMES.get(lvl, f"L{lvl}"),
                })

            return JSONResponse({"results": results})
        except Exception as e:
            logger.error("Maturity search error: %s", e, exc_info=True)
            return JSONResponse({"error": str(e)}, status_code=500)

    def _classification_reason(level: int, agent: int, cli: int, chat: int,
                               completions: int, autonomy: float) -> str:
        """Human-readable explanation based on combined agentic activity (agent_turns + cli_turns)."""
        agentic = agent + cli
        breakdown = f"agent_turns={agent} + cli_turns={cli} = {agentic}"
        if level == 5:
            return (
                f"✅ Matched L5 — Elite Agentic\n"
                f"{breakdown} > 150 (28-day total).\n\n"
                f"This user orchestrates Copilot autonomously at the highest level — "
                f"agent mode and/or CLI are a primary part of their daily workflow."
            )
        if level == 4:
            return (
                f"✅ Matched L4 — Advanced Pilot\n"
                f"{breakdown} in [51, 150] range (28-day total).\n\n"
                f"This user delegates complex tasks to agent mode and/or CLI regularly. "
                f"Strong agentic fluency with significant autonomous AI usage."
            )
        if level == 3:
            return (
                f"✅ Matched L3 — Hybrid Explorer\n"
                f"{breakdown} in [11, 50] range (28-day total).\n\n"
                f"This user has started using agent mode and/or CLI meaningfully. "
                f"The transition phase towards regular agentic adoption."
            )
        if level == 2:
            return (
                f"✅ Matched L2 — Traditionalist\n"
                f"{breakdown} in [1, 10] range (28-day total).\n\n"
                f"This user has tried agent mode and/or CLI but engagement is still limited. "
                f"A candidate for targeted coaching to increase agentic usage."
            )
        return (
            f"✅ Matched L1 — Passive User\n"
            f"{breakdown} = 0 — no agentic activity in 28 days.\n\n"
            f"This user has not used Copilot agent mode or CLI at all. "
            f"Primary target for onboarding, activation campaigns, and peer-coaching."
        )

    @app.get("/api/maturity/user-detail")
    async def api_maturity_user_detail(request: Request, login: str = "", filter: str = "licensed"):
        """Return full 28-day activity breakdown and tier reasoning for a single user."""
        login = login.strip().lower()
        if not login:
            return JSONResponse({"error": "login required"}, status_code=400)
        try:
            user_records = await _get_raw_user_metrics()

            # Collect per-day rows for this user
            agg: dict[str, Any] = {
                "agent": 0, "cli": 0, "chat": 0,
                "completions": 0, "lines_accepted": 0, "dates": set(),
            }
            daily_logs: list[dict[str, Any]] = []

            for rec in user_records:
                rec_login = (
                    rec.get("user_login") or rec.get("github_login") or
                    rec.get("login") or ""
                ).lower()
                if rec_login != login:
                    continue

                day = rec.get("date") or rec.get("day", "")
                if day:
                    agg["dates"].add(day)

                agent_v = rec.get("agent_turns", 0) or 0
                cli_v   = rec.get("cli_turns",   0) or 0
                chat_v  = rec.get("chat_turns",  0) or 0
                comp_v  = rec.get("completions_acceptances", 0) or 0
                lines_v = rec.get("completions_lines_accepted", 0) or 0

                for feat in rec.get("totals_by_feature", []):
                    fn = feat.get("feature", "")
                    if fn == "agent_edit":
                        agent_v += feat.get("code_generation_activity_count", 0)
                    elif fn == "code_completion":
                        comp_v  += feat.get("code_acceptance_activity_count", 0)
                        lines_v += feat.get("loc_added_sum", 0)
                    elif "chat" in fn:
                        chat_v  += feat.get("user_initiated_interaction_count", 0)
                if not comp_v:
                    comp_v  = rec.get("code_acceptance_activity_count", 0) or 0
                    lines_v = lines_v or rec.get("loc_added_sum", 0) or 0
                _cli_data = rec.get("totals_by_cli")
                if isinstance(_cli_data, dict):
                    cli_v += _cli_data.get("session_count", 0)

                agg["agent"] += agent_v; agg["cli"] += cli_v
                agg["chat"]  += chat_v;  agg["completions"] += comp_v
                agg["lines_accepted"] += lines_v

                daily_logs.append({
                    "date": day,
                    "completions": comp_v,
                    "lines_accepted": lines_v,
                    "agent_turns": agent_v,
                    "chat_turns": chat_v,
                    "cli_turns": cli_v,
                })

            daily_logs.sort(key=lambda x: x["date"])
            active_days = len(agg["dates"])
            total_events = agg["agent"] + agg["completions"] + agg["chat"] + agg["cli"]
            autonomy = agg["agent"] / total_events if total_events else 0.0

            level = _classify_user(
                agent=agg["agent"], cli=agg["cli"], chat=agg["chat"],
                completions=agg["completions"], lines_accepted=agg["lines_accepted"],
                active_days=active_days,
            )
            reason = _classification_reason(
                level, agg["agent"], agg["cli"], agg["chat"], agg["completions"], autonomy
            )

            # OrgDB lookup
            db = _get_orgdb()
            org = {"name": "", "surname": "", "email": ""}
            try:
                row = db._conn.execute(
                    "SELECT name, surname, email FROM employees "
                    "WHERE LOWER(github_id) = ? LIMIT 1", (login,)
                ).fetchone()
                if row:
                    org = {
                        "name": row["name"] or "",
                        "surname": row["surname"] or "",
                        "email": row["email"] or "",
                    }
            except Exception:
                pass

            return JSONResponse({
                "login": login,
                "display": " ".join(x for x in [org["name"], org["surname"]] if x) or login,
                "email": org["email"],
                "tier_level": level,
                "tier_name": _LEVEL_NAMES.get(level, f"L{level}"),
                "totals": {
                    "agent_turns":    agg["agent"],
                    "cli_turns":      agg["cli"],
                    "chat_turns":     agg["chat"],
                    "completions":    agg["completions"],
                    "lines_accepted": agg["lines_accepted"],
                    "active_days":    active_days,
                    "autonomy_ratio": round(autonomy, 4),
                },
                "reason": reason,
                "daily_logs": daily_logs,
            })
        except Exception as e:
            logger.error("Maturity user-detail error: %s", e, exc_info=True)
            return JSONResponse({"error": str(e)}, status_code=500)

    @app.get("/api/maturity/loc-by-tier")
    async def api_maturity_loc_by_tier(request: Request, filter: str = "licensed"):
        """Lines of code accepted in the last 7 calendar days, grouped by 28-day maturity tier."""
        try:
            from datetime import date as _date, timedelta as _td

            filter_logins: set[str] | None = _resolve_filter_logins(request)

            # Seat list for licensed filter
            orch = app.state.orchestrator
            licensed: set[str] = set()
            if orch:
                seat_result = await orch._tool_seat_info({"org": config.github_org})
                seats_raw = seat_result.get("seat_info", {}).get("seats", [])
                licensed = {(s.get("login") or "").lower() for s in seats_raw if s.get("login")}

            user_records = await _get_raw_user_metrics()

            # Find latest date in the dataset → last-7-day window
            all_dates = [
                rec.get("date") or rec.get("day", "") for rec in user_records
            ]
            all_dates = [d for d in all_dates if d]
            if not all_dates:
                return JSONResponse({"tiers": [], "total_loc": 0, "week_label": "no data"})

            latest_date = max(all_dates)
            cutoff = (
                _date.fromisoformat(latest_date) - _td(days=6)
            ).isoformat()  # 7-day window: cutoff..latest_date inclusive
            week_label = f"{cutoff} – {latest_date}"

            # 1st pass: full 28-day aggregates per user (for tier classification)
            agg_28d: dict[str, dict[str, Any]] = {}
            # 2nd pass: last-7-day lines_accepted per user
            loc_7d: dict[str, int] = {}

            for rec in user_records:
                login = (
                    rec.get("user_login") or rec.get("github_login") or
                    rec.get("login") or ""
                ).lower()
                if not login:
                    continue
                if filter_logins is not None and login not in filter_logins:
                    continue
                if filter == "licensed" and licensed and login not in licensed:
                    continue

                if login not in agg_28d:
                    agg_28d[login] = {
                        "agent": 0, "cli": 0, "chat": 0,
                        "completions": 0, "lines_accepted": 0, "dates": set(),
                    }
                u = agg_28d[login]
                day = rec.get("date") or rec.get("day", "")
                if day:
                    u["dates"].add(day)

                agent_v = rec.get("agent_turns", 0) or 0
                cli_v   = rec.get("cli_turns",   0) or 0
                chat_v  = rec.get("chat_turns",  0) or 0
                comp_v  = rec.get("completions_acceptances", 0) or 0
                lines_v = rec.get("completions_lines_accepted", 0) or 0
                for feat in rec.get("totals_by_feature", []):
                    fn = feat.get("feature", "")
                    if fn == "agent_edit":
                        agent_v += feat.get("code_generation_activity_count", 0)
                    elif fn == "code_completion":
                        comp_v  += feat.get("code_acceptance_activity_count", 0)
                        lines_v += feat.get("loc_added_sum", 0)
                    elif "chat" in fn:
                        chat_v  += feat.get("user_initiated_interaction_count", 0)
                if not comp_v:
                    comp_v  = rec.get("code_acceptance_activity_count", 0) or 0
                    lines_v = lines_v or rec.get("loc_added_sum", 0) or 0
                _cli_data = rec.get("totals_by_cli")
                if isinstance(_cli_data, dict):
                    cli_v += _cli_data.get("session_count", 0)

                u["agent"] += agent_v; u["cli"] += cli_v
                u["chat"]  += chat_v;  u["completions"] += comp_v
                u["lines_accepted"] += lines_v

                # Accumulate last-7-day LoC
                if day and day >= cutoff:
                    loc_7d[login] = loc_7d.get(login, 0) + lines_v

            # Classify each user by 28-day tier and bucket their 7-day LoC
            tier_loc:   dict[int, int] = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
            tier_users: dict[int, int] = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}

            for login, u in agg_28d.items():
                lvl = _classify_user(
                    agent=u["agent"], cli=u["cli"], chat=u["chat"],
                    completions=u["completions"], lines_accepted=u["lines_accepted"],
                    active_days=len(u["dates"]),
                )
                tier_loc[lvl]   += loc_7d.get(login, 0)
                tier_users[lvl] += 1

            HUMAN_LOC_PER_FTE_WEEK = 50 * 7  # 350 lines/week baseline

            total_loc = sum(tier_loc.values())
            tiers = []
            for lvl in (5, 4, 3, 2, 1):
                loc        = tier_loc[lvl]
                users      = tier_users[lvl]
                virt_fte   = round(loc / HUMAN_LOC_PER_FTE_WEEK, 1)
                add_fte    = round(virt_fte - users, 1)
                tiers.append({
                    "tier_level":      lvl,
                    "tier_name":       _LEVEL_NAMES[lvl],
                    "loc":             loc,
                    "user_count":      users,
                    "loc_pct":         round(loc / total_loc * 100, 1) if total_loc else 0.0,
                    "loc_per_user":    round(loc / users, 1) if users else 0.0,
                    "virtual_fte":     virt_fte,
                    "additional_fte":  add_fte,
                })
            total_virt  = round(total_loc / HUMAN_LOC_PER_FTE_WEEK, 1)
            total_users = sum(tier_users.values())
            return JSONResponse({
                "week_label":         week_label,
                "tiers":              tiers,
                "total_loc":          total_loc,
                "total_virtual_fte":  total_virt,
                "total_additional_fte": round(total_virt - total_users, 1),
            })
        except Exception as e:
            logger.error("Maturity loc-by-tier error: %s", e, exc_info=True)
            return JSONResponse({"error": str(e)}, status_code=500)

    @app.get("/api/maturity/xlsx")
    async def api_maturity_xlsx(request: Request, filter: str = "licensed"):
        """Download an XLSX of all (filtered) users with their maturity tier.

        Columns: name, surname, email, github_login, tier_level, tier_name.
        Joins with OrgDatabase for HR fields when available.
        """
        from datetime import date as _date
        try:
            orch = app.state.orchestrator
            filter_logins: set[str] | None = _resolve_filter_logins(request)

            # 1. Seat list
            licensed: set[str] = set()
            if orch:
                seat_result = await orch._tool_seat_info({"org": config.github_org})
                seats_raw = seat_result.get("seat_info", {}).get("seats", [])
                licensed = {(s.get("login") or "").lower() for s in seats_raw if s.get("login")}

            # 2. Per-user 28-day records — aggregate metrics
            user_records = await _get_raw_user_metrics()
            agg: dict[str, dict[str, Any]] = {}
            for rec in user_records:
                login = (
                    rec.get("user_login") or rec.get("github_login") or
                    rec.get("login") or ""
                ).lower()
                if not login:
                    continue
                if login not in agg:
                    agg[login] = {
                        "agent": 0, "cli": 0, "chat": 0,
                        "completions": 0, "lines_accepted": 0, "dates": set(),
                    }
                u = agg[login]
                day = rec.get("date") or rec.get("day", "")
                if day:
                    u["dates"].add(day)

                agent_v = rec.get("agent_turns", 0) or 0
                cli_v   = rec.get("cli_turns",   0) or 0
                chat_v  = rec.get("chat_turns",  0) or 0
                comp_v  = rec.get("completions_acceptances", 0) or 0
                lines_v = rec.get("completions_lines_accepted", 0) or 0
                for feat in rec.get("totals_by_feature", []):
                    fn = feat.get("feature", "")
                    if fn == "agent_edit":
                        agent_v += feat.get("code_generation_activity_count", 0)
                    elif fn == "code_completion":
                        comp_v  += feat.get("code_acceptance_activity_count", 0)
                        lines_v += feat.get("loc_added_sum", 0)
                    elif "chat" in fn:
                        chat_v  += feat.get("user_initiated_interaction_count", 0)
                if not comp_v:
                    comp_v  = rec.get("code_acceptance_activity_count", 0) or 0
                    lines_v = lines_v or rec.get("loc_added_sum", 0) or 0
                _cli_data = rec.get("totals_by_cli")
                if isinstance(_cli_data, dict):
                    cli_v += _cli_data.get("session_count", 0)
                u["agent"] += agent_v;  u["cli"] += cli_v
                u["chat"]  += chat_v;   u["completions"] += comp_v
                u["lines_accepted"] += lines_v

            # 3. Universe
            if filter == "active":
                universe = set(agg.keys())
            else:
                universe = licensed | set(agg.keys())
            if filter_logins is not None:
                universe = universe & filter_logins

            # 4. OrgDB lookup helper
            db = _get_orgdb()
            def _org_row(login: str) -> dict[str, str]:
                try:
                    row = db._conn.execute(
                        "SELECT name, surname, email FROM employees "
                        "WHERE LOWER(github_id) = ?",
                        (login.lower(),)
                    ).fetchone()
                    if row:
                        return {
                            "name":    row[0] or "",
                            "surname": row[1] or "",
                            "email":   row[2] or "",
                        }
                except Exception:
                    pass
                return {"name": "", "surname": "", "email": ""}

            # 5. Build rows
            rows: list[dict[str, Any]] = []
            for login in sorted(universe):
                u = agg.get(login, {})
                lvl = _classify_user(
                    agent=u.get("agent", 0),
                    cli=u.get("cli", 0),
                    chat=u.get("chat", 0),
                    completions=u.get("completions", 0),
                    lines_accepted=u.get("lines_accepted", 0),
                    active_days=len(u.get("dates", set())),
                )
                org = _org_row(login)
                rows.append({
                    "name":       org["name"],
                    "surname":    org["surname"],
                    "email":      org["email"],
                    "github_login": login,
                    "tier_level": lvl,
                    "tier_name":  _LEVEL_NAMES.get(lvl, f"L{lvl}"),
                })

            # 6. Build XLSX
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment

            TIER_COLORS = {
                5: "1C3A1C",  # dark green bg → amber text N/A: use cell fill
                4: "1C3A1C",
                3: "0D2137",
                2: "2A2A2A",
                1: "2A0D0D",
            }
            TIER_FONT_COLORS = {
                5: "F59E0B",
                4: "3FB950",
                3: "58A6FF",
                2: "8B949E",
                1: "F85149",
            }

            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Maturity Tiers"

            header_cols = ["Name", "Surname", "Email", "GitHub Login", "Tier Level", "Tier Name"]
            header_fill = PatternFill("solid", fgColor="0D1117")
            header_font = Font(bold=True, color="58A6FF")
            for col_idx, h in enumerate(header_cols, 1):
                cell = ws.cell(row=1, column=col_idx, value=h)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            col_widths = [18, 18, 32, 22, 12, 22]
            for i, w in enumerate(col_widths, 1):
                ws.column_dimensions[
                    openpyxl.utils.get_column_letter(i)
                ].width = w

            for r_idx, row in enumerate(rows, 2):
                lvl = row["tier_level"]
                fg = TIER_FONT_COLORS.get(lvl, "FFFFFF")
                ws.cell(row=r_idx, column=1, value=row["name"])
                ws.cell(row=r_idx, column=2, value=row["surname"])
                ws.cell(row=r_idx, column=3, value=row["email"])
                ws.cell(row=r_idx, column=4, value=row["github_login"])
                tier_level_cell = ws.cell(row=r_idx, column=5, value=f"L{lvl}")
                tier_level_cell.font = Font(bold=True, color=fg)
                tier_level_cell.alignment = Alignment(horizontal="center")
                tier_name_cell = ws.cell(row=r_idx, column=6, value=row["tier_name"])
                tier_name_cell.font = Font(color=fg)

            buf = io.BytesIO()
            wb.save(buf)
            buf.seek(0)

            filename = f"maturity-tiers-{_date.today().isoformat()}.xlsx"
            return StreamingResponse(
                buf,
                media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                headers={"Content-Disposition": f"attachment; filename={filename}"},
            )
        except Exception as e:
            logger.error("Maturity XLSX error: %s", e, exc_info=True)
            return JSONResponse({"error": str(e)}, status_code=500)

    @app.get("/api/maturity/trend")
    async def api_maturity_trend(request: Request):
        """Return weekly maturity band % using the same per-user classification as the pyramid.

        Splits the 28-day user records into 7-day buckets and classifies each user
        per week with proportionally scaled thresholds (28-day ÷ 4).
        Denominator = licensed seats (consistent with the pyramid Licensed view).
        """
        try:
            _trend_key = _filter_cache_key(request)
            _cached_trend = _maturity_store.get(f"maturity:trend:{_trend_key}")
            if _cached_trend is not None:
                logger.debug("Maturity trend: cache hit (%s)", _trend_key)
                return JSONResponse(_cached_trend)

            # 1. Per-user 28-day records (same source as pyramid)
            user_records = await _get_raw_user_metrics()

            # 2. Seat count for denominator
            total_seats = 0
            orch = app.state.orchestrator
            if orch:
                seat_result = await orch._tool_seat_info({"org": config.github_org})
                seats = seat_result.get("seat_info", {}).get("seats", [])
                total_seats = len(seats)

            # 3. Combined filter (org hierarchy + attribute filters)
            filter_logins: set[str] | None = _resolve_filter_logins(request)
            if filter_logins is not None:
                user_records = [
                    r for r in user_records
                    if (r.get("user_login") or r.get("github_login") or "").lower() in filter_logins
                ]
                total_seats = len(filter_logins) or total_seats

            # 4. Determine all dates present and build week buckets (newest first → reverse)
            from datetime import date as _date, timedelta as _td
            all_dates: set[str] = {
                r.get("date") or r.get("day", "") for r in user_records
                if r.get("date") or r.get("day")
            }
            if not all_dates:
                return JSONResponse({"weeks": [], "l5_l4_pct": [], "l3_pct": [], "l1_l2_pct": []})

            max_date = max(all_dates)
            try:
                anchor = _date.fromisoformat(max_date)
            except ValueError:
                anchor = _date.today()

            # Build 4 weekly buckets covering the 28-day window
            # Week 1 = oldest (days 22–28), Week 4 = newest (days 1–7)
            week_buckets: list[tuple[_date, _date]] = []
            for i in range(3, -1, -1):
                end   = anchor - _td(days=i * 7)
                start = end - _td(days=6)
                week_buckets.append((start, end))

            # Weekly thresholds = 28-day ÷ 4 (agent_turns + cli_turns)
            # L5: >=38  |  L4: >=13  |  L3: >=3  |  L2: >=1  |  L1: 0
            def _classify_week(agent: int, cli: int, chat: int,
                               completions: int, lines: int) -> int:
                agentic = agent + cli
                if agentic >= 38:  return 5
                if agentic >= 13:  return 4
                if agentic >= 3:   return 3
                if agentic >= 1:   return 2
                return 1

            labels: list[str] = []
            l5_l4: list[float] = []
            l3: list[float] = []
            l1_l2: list[float] = []

            for wk_start, wk_end in week_buckets:
                wk_start_s = wk_start.isoformat()
                wk_end_s   = wk_end.isoformat()

                # Aggregate per-user for this week
                wk_agg: dict[str, dict[str, Any]] = {}
                for rec in user_records:
                    d = rec.get("date") or rec.get("day", "")
                    if not d or not (wk_start_s <= d <= wk_end_s):
                        continue
                    login = (
                        rec.get("user_login") or rec.get("github_login") or
                        rec.get("login") or ""
                    ).lower()
                    if not login:
                        continue
                    if login not in wk_agg:
                        wk_agg[login] = {
                            "agent": 0, "cli": 0, "chat": 0, "completions": 0, "lines": 0
                        }
                    u = wk_agg[login]

                    agent_v = rec.get("agent_turns", 0) or 0
                    cli_v   = rec.get("cli_turns",   0) or 0
                    chat_v  = rec.get("chat_turns",  0) or 0
                    comp_v  = rec.get("completions_acceptances", 0) or 0
                    lines_v = rec.get("completions_lines_accepted", 0) or 0

                    for feat in rec.get("totals_by_feature", []):
                        fn = feat.get("feature", "")
                        if fn == "agent_edit":
                            agent_v += feat.get("code_generation_activity_count", 0)
                        elif fn == "code_completion":
                            comp_v  += feat.get("code_acceptance_activity_count", 0)
                            lines_v += feat.get("loc_added_sum", 0)
                        elif "chat" in fn:
                            chat_v  += feat.get("user_initiated_interaction_count", 0)
                    if not comp_v:
                        comp_v  = rec.get("code_acceptance_activity_count", 0) or 0
                        lines_v = lines_v or rec.get("loc_added_sum", 0) or 0

                    _cli_data = rec.get("totals_by_cli")
                    if isinstance(_cli_data, dict):
                        cli_v += _cli_data.get("session_count", 0)

                    u["agent"] += agent_v
                    u["cli"]   += cli_v
                    u["chat"]  += chat_v
                    u["completions"] += comp_v
                    u["lines"] += lines_v

                dist: dict[int, int] = {1: 0, 2: 0, 3: 0, 4: 0, 5: 0}
                for u in wk_agg.values():
                    lvl = _classify_week(
                        u["agent"], u["cli"], u["chat"], u["completions"], u["lines"]
                    )
                    dist[lvl] += 1

                denom = total_seats or len(wk_agg) or 1
                l5_l4_pct  = round((dist[5] + dist[4]) / denom * 100, 1)
                l3_pct_val = round(dist[3] / denom * 100, 1)
                l1_l2_pct  = round((dist[1] + dist[2]) / denom * 100, 1)

                label = f"W{wk_end.isocalendar()[1]} ({wk_end.strftime('%d/%m')})"
                labels.append(label)
                l5_l4.append(l5_l4_pct)
                l3.append(l3_pct_val)
                l1_l2.append(l1_l2_pct)

            trend_result = {
                "weeks": labels,
                "l5_l4_pct": l5_l4,
                "l3_pct": l3,
                "l1_l2_pct": l1_l2,
            }
            _maturity_store.set(f"maturity:trend:{_trend_key}", trend_result)
            logger.info("Maturity trend cached for 24 h (key=%s)", _trend_key)
            return JSONResponse(trend_result)
        except Exception as e:
            logger.error("Maturity trend error: %s", e, exc_info=True)
            return JSONResponse({"error": str(e)}, status_code=500)

    @app.on_event("startup")
    async def _prewarm_caches() -> None:
        """Pre-warm the weekly agent series cache in the background on startup.

        This fires the 13 API calls once at startup so the first real user
        request hits a warm cache instead of experiencing a multi-second delay.
        The task runs in the background and never blocks server startup.
        """
        async def _warm() -> None:
            try:
                from starlette.datastructures import QueryParams

                class _StubRequest:
                    """Minimal stand-in for a Starlette Request with no query params."""
                    query_params = QueryParams("")

                await _get_weekly_agent_series(_StubRequest())  # type: ignore[arg-type]
                logger.info("Cache pre-warm: weekly agent series ready")
            except Exception as exc:
                logger.warning("Cache pre-warm failed (non-fatal): %s", exc)

        asyncio.create_task(_warm())

    return app
