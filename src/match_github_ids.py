"""
match_github_ids.py

Two-pass matching of GitHub usernames from the Copilot license file into the
org users file.

Pass 1 – exact email match (case-insensitive):
    Compare 'Email - Primary Work' (col 39, org file) vs 'Mail' (Copilot file).

Pass 2 – name-from-email fallback:
    Parse FirstName / LastName out of the Copilot email local part,
    normalise accents and split CamelCase names, then match against
    'Name' (col 12) + 'Surname' (col 13) in the org file.
    Only unambiguous 1-to-1 matches are accepted.

Outputs:
    org_users_mail_updated.xlsx  – org file with github_id column filled
    conflict_report.txt          – records where existing github_id differs
                                   from the Copilot file value
"""

import re
import unicodedata
from pathlib import Path

import openpyxl

BASE_DIR = Path(__file__).resolve().parent.parent
ORG_FILE = BASE_DIR / "org_users_mail.xlsx"
COPILOT_FILE = BASE_DIR / "GH_Copilot_Aprile26 1.xlsx"
OUTPUT_FILE = BASE_DIR / "org_users_mail_updated.xlsx"
CONFLICT_REPORT = BASE_DIR / "conflict_report.txt"

# Column indices (0-based) in org file
COL_NAME = 12
COL_SURNAME = 13
COL_GITHUB_ID = 15
COL_PRIMARY_EMAIL = 39


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def strip_accents(s: str) -> str:
    """Remove diacritics from a string."""
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfkd if unicodedata.category(c) != "Mn")


def normalise(s: str) -> str:
    """Uppercase + strip accents + collapse whitespace."""
    if not s:
        return ""
    return " ".join(strip_accents(s).upper().split())


def split_camel(s: str) -> list[str]:
    """Split 'AdelinaCosmina' → ['ADELINA', 'COSMINA']."""
    parts = re.sub(r"([A-Z][a-z]+)", r" \1", s).split()
    return [p.upper() for p in parts if p]


def email_to_name_parts(email_local: str) -> tuple[list[str], str]:
    """
    Parse email local part into (first_name_tokens, last_name).

    'AdelinaCosmina.Darie'  → (['ADELINA', 'COSMINA'], 'DARIE')
    'Adolfo.Volpe'          → (['ADOLFO'], 'VOLPE')
    'first.middle.last'     → (['FIRST', 'MIDDLE'], 'LAST')
    """
    parts = email_local.split(".")
    raw_first_tokens = parts[:-1]
    raw_last = parts[-1]

    first_tokens: list[str] = []
    for token in raw_first_tokens:
        camel = split_camel(token)
        if camel:
            first_tokens.extend(camel)
        else:
            first_tokens.append(normalise(token))

    last_name = normalise(raw_last)
    return first_tokens, last_name


# ---------------------------------------------------------------------------
# Loading data
# ---------------------------------------------------------------------------

def load_copilot_file() -> dict[str, str]:
    """Return {email_lower: github_username} from the Copilot file."""
    wb = openpyxl.load_workbook(COPILOT_FILE, read_only=True)
    ws = wb.active
    result: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        username, mail, *_ = row
        if mail and username:
            result[str(mail).lower().strip()] = str(username).strip()
    wb.close()
    return result


# ---------------------------------------------------------------------------
# Name index for org file (for Pass 2)
# ---------------------------------------------------------------------------

def build_name_index(rows: list[tuple]) -> dict[tuple, list[int]]:
    """
    Map (first_name_normalised, surname_normalised) → [row_indices].
    Rows with compound first names (e.g. 'ADELINA COSMINA') are stored
    under every sub-key so CamelCase email tokens can find them.
    """
    index: dict[tuple, list[int]] = {}

    for idx, row in enumerate(rows):
        name_raw = row[COL_NAME]
        surname_raw = row[COL_SURNAME]
        if not name_raw or not surname_raw:
            continue
        name_norm = normalise(str(name_raw))
        surname_norm = normalise(str(surname_raw))

        # Full compound name as-is  (e.g. 'ADELINA COSMINA')
        key = (name_norm, surname_norm)
        index.setdefault(key, []).append(idx)

        # Each individual token of a compound first name
        tokens = name_norm.split()
        if len(tokens) > 1:
            for token in tokens:
                key2 = (token, surname_norm)
                index.setdefault(key2, []).append(idx)

    return index


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main() -> None:
    print(f"Loading org file:     {ORG_FILE}")
    print(f"Loading Copilot file: {COPILOT_FILE}")

    copilot_map = load_copilot_file()
    print(f"  Copilot users loaded: {len(copilot_map)}")

    wb = openpyxl.load_workbook(ORG_FILE)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    # Pre-build name index for Pass 2
    name_index = build_name_index(data_rows)

    # Build email index for Pass 1
    # email_lower → row_index (0-based in data_rows)
    email_to_idx: dict[str, int] = {}
    for idx, row in enumerate(data_rows):
        org_email = row[COL_PRIMARY_EMAIL]
        if org_email:
            email_to_idx[str(org_email).lower().strip()] = idx

    # Track outcomes
    matched_pass1 = 0
    matched_pass2 = 0
    skipped_same = 0
    conflicts: list[str] = []
    unmatched_copilot = 0

    # We will write updates back as a dict: row_index -> new_github_id
    updates: dict[int, str] = {}

    # -----------------------------------------------------------------------
    # Pass 1 – email match
    # -----------------------------------------------------------------------
    for copilot_email, username in copilot_map.items():
        if copilot_email not in email_to_idx:
            unmatched_copilot += 1
            continue

        idx = email_to_idx[copilot_email]
        row = data_rows[idx]
        existing = row[COL_GITHUB_ID]

        if existing:
            existing_str = str(existing).strip()
            if existing_str.lower() == username.lower():
                skipped_same += 1
            else:
                # Conflict: different value already present
                conflicts.append(
                    f"  {row[COL_NAME]} {row[COL_SURNAME]}"
                    f" | org_email: {row[COL_PRIMARY_EMAIL]}"
                    f" | existing github_id: {existing_str}"
                    f" | copilot username: {username}"
                )
        else:
            updates[idx] = username
            matched_pass1 += 1

    print(f"\nPass 1 (email match):  {matched_pass1} new matches")
    print(f"  Already had same ID: {skipped_same}")
    print(f"  Conflicts:           {len(conflicts)}")
    print(f"  Not in org file:     {unmatched_copilot}")

    # -----------------------------------------------------------------------
    # Pass 2 – name-from-email fallback
    # -----------------------------------------------------------------------
    # Build a set of org row indices already resolved (by email or updates)
    resolved_indices: set[int] = set(updates.keys())
    for idx, row in enumerate(data_rows):
        if row[COL_GITHUB_ID]:
            resolved_indices.add(idx)

    # Build reverse map: org_email_lower → copilot username (already done above)
    # For pass 2 we need: copilot_email → username for emails NOT matched in pass 1
    pass2_candidates: list[tuple[str, str]] = [
        (email, user)
        for email, user in copilot_map.items()
        if email not in email_to_idx
    ]

    pass2_ambiguous = 0
    for copilot_email, username in pass2_candidates:
        local = copilot_email.split("@")[0]
        first_tokens, last_name = email_to_name_parts(local)
        if not first_tokens or not last_name:
            continue

        # Try the full compound first name first, then individual tokens
        candidate_sets: list[tuple] = []
        full_first = " ".join(first_tokens)
        candidate_sets.append((full_first, last_name))
        for token in first_tokens:
            candidate_sets.append((token, last_name))

        found_indices: set[int] = set()
        for key in candidate_sets:
            found_indices.update(name_index.get(key, []))

        # Filter out already-resolved rows
        eligible = [i for i in found_indices if i not in resolved_indices]
        if len(eligible) == 1:
            idx = eligible[0]
            updates[idx] = username
            resolved_indices.add(idx)
            matched_pass2 += 1
        elif len(eligible) > 1:
            pass2_ambiguous += 1

    print(f"\nPass 2 (name fallback): {matched_pass2} new matches")
    print(f"  Ambiguous (skipped):  {pass2_ambiguous}")

    # -----------------------------------------------------------------------
    # Write updates back to the workbook
    # -----------------------------------------------------------------------
    # Find the actual column letter for github_id (col index 15, 1-based = 16)
    github_col = COL_GITHUB_ID + 1  # openpyxl is 1-based

    for row_idx, username in updates.items():
        excel_row = row_idx + 2  # +1 for header, +1 for 1-based
        ws.cell(row=excel_row, column=github_col, value=username)

    wb.save(OUTPUT_FILE)
    print(f"\nSaved updated file to: {OUTPUT_FILE}")

    total_new = matched_pass1 + matched_pass2
    print(f"\n=== Summary ===")
    print(f"  Total new github_id entries written: {total_new}")
    print(f"  Skipped (already had correct ID):    {skipped_same}")
    print(f"  Conflicts (NOT overwritten):         {len(conflicts)}")
    print(f"  Copilot users with no org match:     {unmatched_copilot}")

    # -----------------------------------------------------------------------
    # Write conflict report
    # -----------------------------------------------------------------------
    with open(CONFLICT_REPORT, "w", encoding="utf-8") as f:
        f.write("=== github_id Conflict Report ===\n")
        f.write(
            "These records already had a github_id that DIFFERS from the Copilot file.\n"
            "They were NOT overwritten. Please resolve manually.\n\n"
        )
        if conflicts:
            f.write("\n".join(conflicts) + "\n")
        else:
            f.write("No conflicts found.\n")

    print(f"\nConflict report:       {CONFLICT_REPORT}")
    if conflicts:
        print("\nConflicting records:")
        for c in conflicts:
            print(c)


if __name__ == "__main__":
    main()
