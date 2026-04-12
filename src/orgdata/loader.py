"""Load organizational structure data from Excel files."""

from __future__ import annotations

import logging
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl

from .models import Employee

try:
    import xlrd  # optional – only needed for legacy .xls files
    _XLRD_AVAILABLE = True
except ImportError:
    _XLRD_AVAILABLE = False

logger = logging.getLogger(__name__)

# Normalized header name (lowercase, stripped) → Employee field name
_HEADER_TO_FIELD: dict[str, str] = {
    "company ref id": "company_ref_id",
    "company": "company",
    "location country": "location_country",
    "supervisory organization": "supervisory_org",
    "supervisory organization id": "supervisory_org_id",
    "employee id": "employee_id",
    "position title": "position_title",
    "matricola": "matricola",
    "worker type": "worker_type",
    "name": "name",
    "surname": "surname",
    "email": "email",
    "github_id": "github_id",
    "location": "location",
    "location id": "location_id",
    "cdc code": "cdc_code",
    "cost center - id": "cost_center_id",
    "is manager": "is_manager",
    "hr business partner": "hr_business_partner",
    "hr director": "hr_director",
    "business title": "business_title",
    "job profile": "job_profile",
    "job family": "job_family",
    "employee type": "employee_type",
    "contract type": "contract_type",
    "management level": "management_level",
    "time type": "time_type",
    "fte": "fte",
    "original hire date": "original_hire_date",
    "continuous service date": "continuous_service_date",
    "length of service": "length_of_service",
    "company service date": "company_service_date",
    "age": "age",
    "time in position": "time_in_position",
    "email - primary work": "email_work",
    "job code": "job_code",
    "job title": "job_title",
    "job level": "job_level",
    "job category": "job_category",
    "gender": "gender",
    "date of birth": "date_of_birth",
    "top level sup org": "top_level_sup_org",
    "sup org level 2": "sup_org_level_2",
    "sup org level 3": "sup_org_level_3",
    "sup org level 4": "sup_org_level_4",
    "sup org level 5": "sup_org_level_5",
    "sup org level 6": "sup_org_level_6",
    "sup org level 7": "sup_org_level_7",
    "sup org level 8": "sup_org_level_8",
    "sup org level 9": "sup_org_level_9",
    "sup org level 10": "sup_org_level_10",
}

# Fields required for every import — upload is blocked if any are absent
REQUIRED_FIELDS: frozenset[str] = frozenset({
    "matricola", "name", "surname", "business_title", "job_profile", "job_family", "email",
})


def _normalize_header(h: Any) -> str:
    if h is None:
        return ""
    return str(h).strip().lower()


def _lookup_field(norm: str) -> str | None:
    """Look up an Employee field name from a normalized header string.

    Tries exact match first, then underscore→space substitution so that
    headers like ``job_profile`` match the ``"job profile"`` key.
    """
    return _HEADER_TO_FIELD.get(norm) or _HEADER_TO_FIELD.get(norm.replace("_", " "))


def _safe_str(val: Any) -> str:
    if val is None:
        return ""
    # xlrd returns integers/IDs as float (e.g. 12345.0) — strip the .0
    if isinstance(val, float) and val == int(val):
        return str(int(val))
    return str(val).strip()


def _safe_date(val: Any) -> date | None:
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    return None


def _safe_int(val: Any) -> int | None:
    if val is None:
        return None
    try:
        return int(val)
    except (ValueError, TypeError):
        return None


def _safe_float(val: Any) -> float:
    if val is None:
        return 1.0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 1.0


def _read_all_rows(file_path: Path) -> list[tuple]:
    """Read all rows from an .xlsx or legacy .xls file."""
    if file_path.suffix.lower() == ".xls":
        if not _XLRD_AVAILABLE:
            raise ImportError(
                "The 'xlrd' package is required to read legacy .xls files. "
                "Install it with: pip install xlrd>=2.0"
            )
        wb = xlrd.open_workbook(str(file_path))
        ws = wb.sheet_by_name("Data") if "Data" in wb.sheet_names() else wb.sheet_by_index(0)
        rows: list[tuple] = []
        for i in range(ws.nrows):
            row_vals: list[Any] = []
            for j in range(ws.ncols):
                cell = ws.cell(i, j)
                # xlrd type 3 = XL_CELL_DATE
                if cell.ctype == 3:
                    try:
                        dt = xlrd.xldate_as_datetime(cell.value, wb.datemode)
                        row_vals.append(dt.date())
                    except Exception:
                        row_vals.append(cell.value)
                else:
                    row_vals.append(cell.value if cell.value != "" else None)
            rows.append(tuple(row_vals))
        return rows
    else:
        wb = openpyxl.load_workbook(str(file_path), read_only=True, data_only=True)
        ws = wb["Data"] if "Data" in wb.sheetnames else wb.worksheets[0]
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
        return rows


def parse_xls_preview(file_path: str | Path) -> dict[str, Any]:
    """Read XLS headers and first 10 data rows for the column-selection overlay.

    Returns a dict with:
    - ``columns``: list of column descriptors — each has ``header`` (original
      string), ``field`` (Employee field name or None), ``required`` (bool),
      ``status`` ("required" | "mapped" | "unmapped"), ``col_idx`` (int).
      Duplicate headers (e.g. two "Age" columns) are deduplicated; the last
      occurrence wins.
    - ``sample_rows``: up to 10 data rows as lists of strings, aligned to
      ``columns`` order.
    - ``missing_required``: required field names not found in the file.
    """
    file_path = Path(file_path)
    raw = _read_all_rows(file_path)[:12]

    if not raw:
        return {"columns": [], "sample_rows": [], "missing_required": sorted(REQUIRED_FIELDS)}

    header_row = raw[0]
    data_rows = [r for r in raw[1:] if any(v is not None for v in r[:20])][:10]

    # Walk headers left-to-right; last occurrence of any duplicate wins.
    seen: dict[str, tuple[str | None, int]] = {}  # norm → (field, col_idx)
    order: list[str] = []                          # original headers in first-seen order

    for idx, h in enumerate(header_row):
        norm = _normalize_header(h)
        if not norm:
            continue
        original = str(h).strip()
        field = _lookup_field(norm)
        if norm not in seen:
            order.append(original)
        seen[norm] = (field, idx)

    columns: list[dict[str, Any]] = []
    for original in order:
        norm = _normalize_header(original)
        field, col_idx = seen[norm]
        required = field in REQUIRED_FIELDS if field else False
        if field is None:
            status = "unmapped"
        elif required:
            status = "required"
        else:
            status = "mapped"
        columns.append(
            {"header": original, "field": field, "required": required,
             "status": status, "col_idx": col_idx}
        )

    # Build sample rows
    sample_rows: list[list[str]] = []
    for row in data_rows:
        values: list[str] = []
        for col in columns:
            idx = col["col_idx"]
            val = row[idx] if idx < len(row) else None
            if isinstance(val, datetime):
                val = str(val.date())
            elif isinstance(val, date):
                val = str(val)
            elif val is not None:
                val = str(val)
            values.append(val or "")
        sample_rows.append(values)

    found_fields = {c["field"] for c in columns if c["field"]}
    missing_required = sorted(REQUIRED_FIELDS - found_fields)

    return {"columns": columns, "sample_rows": sample_rows, "missing_required": missing_required}


def load_with_selection(
    file_path: str | Path,
    selected_headers: list[str] | None = None,
) -> list[Employee]:
    """Load employees from an Excel file, importing only selected columns.

    Args:
        file_path: Path to the Excel file.
        selected_headers: List of original header strings the user selected.
            If ``None``, all mappable headers are included.

    Returns:
        List of :class:`Employee` instances.
    """
    file_path = Path(file_path)
    all_rows = _read_all_rows(file_path)
    if not all_rows:
        return []

    header_row = all_rows[0]
    data_rows = iter(all_rows[1:])

    selected_norm: set[str] | None = (
        {_normalize_header(h) for h in selected_headers}
        if selected_headers is not None
        else None
    )

    # field → col_idx; last occurrence wins for duplicate headers
    field_to_idx: dict[str, int] = {}
    for idx, h in enumerate(header_row):
        norm = _normalize_header(h)
        if not norm:
            continue
        if selected_norm is not None and norm not in selected_norm:
            continue
        field = _lookup_field(norm)
        if field:
            field_to_idx[field] = idx

    def _get(row: tuple, field: str, default: Any = None) -> Any:
        idx = field_to_idx.get(field)
        if idx is None or idx >= len(row):
            return default
        return row[idx]

    employees: list[Employee] = []
    for row in data_rows:
        if not any(v is not None for v in row[:20]):
            continue

        employee_id = _safe_str(_get(row, "employee_id")) or _safe_str(_get(row, "matricola"))
        if not employee_id:
            continue

        is_manager_raw = _safe_str(_get(row, "is_manager"))
        is_manager: bool | None = (
            True if is_manager_raw.lower() in ("yes", "true", "1", "y")
            else (False if is_manager_raw else None)
        )

        emp = Employee(
            employee_id=employee_id,
            matricola=_safe_str(_get(row, "matricola")),
            github_id=_safe_str(_get(row, "github_id")) or None,
            name=_safe_str(_get(row, "name")),
            surname=_safe_str(_get(row, "surname")),
            email=_safe_str(_get(row, "email")) or None,
            email_work=_safe_str(_get(row, "email_work")) or None,
            gender=_safe_str(_get(row, "gender")) or None,
            age=_safe_int(_get(row, "age")),
            date_of_birth=_safe_date(_get(row, "date_of_birth")),
            company=_safe_str(_get(row, "company")),
            company_ref_id=_safe_str(_get(row, "company_ref_id")),
            location_country=_safe_str(_get(row, "location_country")),
            location=_safe_str(_get(row, "location")),
            location_id=_safe_str(_get(row, "location_id")),
            position_title=_safe_str(_get(row, "position_title")),
            business_title=_safe_str(_get(row, "business_title")),
            job_profile=_safe_str(_get(row, "job_profile")),
            job_family=_safe_str(_get(row, "job_family")),
            job_code=_safe_str(_get(row, "job_code")),
            job_title=_safe_str(_get(row, "job_title")),
            job_level=_safe_str(_get(row, "job_level")),
            job_category=_safe_str(_get(row, "job_category")),
            management_level=_safe_str(_get(row, "management_level")),
            worker_type=_safe_str(_get(row, "worker_type")),
            employee_type=_safe_str(_get(row, "employee_type")),
            contract_type=_safe_str(_get(row, "contract_type")),
            time_type=_safe_str(_get(row, "time_type")),
            fte=_safe_float(_get(row, "fte")),
            is_manager=is_manager,
            original_hire_date=_safe_date(_get(row, "original_hire_date")),
            continuous_service_date=_safe_date(_get(row, "continuous_service_date")),
            company_service_date=_safe_date(_get(row, "company_service_date")),
            length_of_service=_safe_str(_get(row, "length_of_service")),
            time_in_position=_safe_str(_get(row, "time_in_position")),
            cdc_code=_safe_str(_get(row, "cdc_code")),
            cost_center_id=_safe_str(_get(row, "cost_center_id")),
            supervisory_org=_safe_str(_get(row, "supervisory_org")),
            supervisory_org_id=_safe_str(_get(row, "supervisory_org_id")),
            top_level_sup_org=_safe_str(_get(row, "top_level_sup_org")),
            sup_org_level_2=_safe_str(_get(row, "sup_org_level_2")),
            sup_org_level_3=_safe_str(_get(row, "sup_org_level_3")),
            sup_org_level_4=_safe_str(_get(row, "sup_org_level_4")),
            sup_org_level_5=_safe_str(_get(row, "sup_org_level_5")),
            sup_org_level_6=_safe_str(_get(row, "sup_org_level_6")),
            sup_org_level_7=_safe_str(_get(row, "sup_org_level_7")),
            sup_org_level_8=_safe_str(_get(row, "sup_org_level_8")),
            sup_org_level_9=_safe_str(_get(row, "sup_org_level_9")),
            sup_org_level_10=_safe_str(_get(row, "sup_org_level_10")),
            hr_business_partner=_safe_str(_get(row, "hr_business_partner")),
            hr_director=_safe_str(_get(row, "hr_director")),
        )
        employees.append(emp)

    logger.info("Loaded %d employees from %s", len(employees), file_path.name)
    return employees


class OrgDataLoader:
    """Load employee org structure from an Excel file.

    Args:
        file_path: Path to the Excel (.xlsx) file.
    """

    def __init__(self, file_path: str | Path) -> None:
        self.file_path = Path(file_path)
        if not self.file_path.exists():
            raise FileNotFoundError(f"Org structure file not found: {self.file_path}")

    def load(self, selected_headers: list[str] | None = None) -> list[Employee]:
        """Load employee records from the Excel file.

        Args:
            selected_headers: Optional subset of header names to include.
                If ``None``, all recognised headers are imported.

        Returns:
            List of :class:`Employee` instances.
        """
        return load_with_selection(self.file_path, selected_headers)
